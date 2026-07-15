#!/usr/bin/env python3
"""
Hero eCatalogue URL Migration Bug Fix Verification
Tests the fix for Hero eCatalogue integration after URL migration from 
ecatalogue.heromotocorp.com to ecatalogue.heromotocorp.biz:8080
"""

import requests
import io
import sys
from openpyxl import Workbook

# Backend URL
BASE_URL = "https://008a5671-c16f-4b24-91b3-151477b7ed8b.preview.emergentagent.com/api"

# Test credentials from /app/memory/test_credentials.md
OWNER_USERNAME = "admin"
OWNER_PASSWORD = "admin123"

# Global state
owner_token = None

def print_test(name):
    """Print test name"""
    print(f"\n{'='*80}")
    print(f"TEST: {name}")
    print('='*80)

def print_result(passed, message=""):
    """Print test result"""
    status = "✅ PASS" if passed else "❌ FAIL"
    print(f"{status}: {message}")
    return passed

def print_response_details(response):
    """Print detailed response information"""
    print(f"Status Code: {response.status_code}")
    print(f"Response Headers: {dict(response.headers)}")
    try:
        print(f"Response Body: {response.json()}")
    except:
        print(f"Response Text: {response.text[:500]}")

def create_test_inventory_xlsx():
    """Create a minimal test inventory Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Headers
    ws['A1'] = "Part No"
    ws['B1'] = "Stock Qty"
    
    # Test data - Hero part
    ws['A2'] = "23121KST901"
    ws['B2'] = 10
    
    # Test data - TVS part for regression test
    ws['A3'] = "N3012050"
    ws['B3'] = 5
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def test_1_login():
    """Test 1: Login with admin/admin123 and save access_token"""
    global owner_token
    
    print_test("1. Login - POST /api/auth/login")
    
    try:
        response = requests.post(
            f"{BASE_URL}/auth/login",
            json={"username": OWNER_USERNAME, "password": OWNER_PASSWORD},
            timeout=30
        )
        
        print_response_details(response)
        
        if response.status_code != 200:
            return print_result(False, f"Login failed with status {response.status_code}")
        
        data = response.json()
        
        if "access_token" not in data:
            return print_result(False, "Missing access_token in response")
        
        owner_token = data["access_token"]
        
        return print_result(True, f"Login successful, token obtained: {owner_token[:20]}...")
        
    except Exception as e:
        return print_result(False, f"Exception during login: {str(e)}")

def test_2_check_inventory_freshness():
    """Test 2: Check inventory freshness and upload if needed"""
    global owner_token
    
    print_test("2. Check Inventory Freshness - GET /api/dashboard/stats?system=hero")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        response = requests.get(
            f"{BASE_URL}/dashboard/stats?system=hero",
            headers=headers,
            timeout=30
        )
        
        print_response_details(response)
        
        if response.status_code != 200:
            return print_result(False, f"Dashboard stats failed with status {response.status_code}")
        
        data = response.json()
        inventory_status = data.get("inventory_status", {})
        
        never_uploaded = inventory_status.get("never_uploaded", False)
        is_stale = inventory_status.get("is_stale", False)
        
        print(f"Inventory Status: never_uploaded={never_uploaded}, is_stale={is_stale}")
        
        if never_uploaded or is_stale:
            print("\n⚠️  Inventory is stale or never uploaded. Uploading fresh inventory...")
            return upload_inventory()
        else:
            return print_result(True, "Inventory is fresh, no upload needed")
            
    except Exception as e:
        return print_result(False, f"Exception checking inventory: {str(e)}")

def upload_inventory():
    """Upload test inventory"""
    global owner_token
    
    print_test("2b. Upload Inventory - POST /api/inventory/upload")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        # Create test Excel file
        xlsx_bytes = create_test_inventory_xlsx()
        
        files = {
            'file': ('test_inventory.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        data = {
            'part_no': 'Part No',
            'stock_qty': 'Stock Qty',
            'replace': 'true'
        }
        
        response = requests.post(
            f"{BASE_URL}/inventory/upload",
            headers=headers,
            files=files,
            data=data,
            timeout=30
        )
        
        print_response_details(response)
        
        if response.status_code != 200:
            return print_result(False, f"Inventory upload failed with status {response.status_code}")
        
        result = response.json()
        
        if not result.get("success"):
            return print_result(False, f"Upload not successful: {result}")
        
        imported = result.get("imported", 0)
        if imported < 1:
            return print_result(False, f"No items imported: {result}")
        
        return print_result(True, f"Inventory uploaded successfully, imported {imported} items")
        
    except Exception as e:
        return print_result(False, f"Exception during inventory upload: {str(e)}")

def test_3_hero_search_primary():
    """Test 3: Hero search happy path - GET /api/hero/search?q=23121KST901"""
    global owner_token
    
    print_test("3. Hero Search Happy Path - GET /api/hero/search?q=23121KST901")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        response = requests.get(
            f"{BASE_URL}/hero/search?q=23121KST901",
            headers=headers,
            timeout=30
        )
        
        print_response_details(response)
        
        if response.status_code != 200:
            return print_result(False, f"Hero search failed with status {response.status_code}: {response.text}")
        
        data = response.json()
        
        # Check parts array exists
        if "parts" not in data:
            return print_result(False, "Missing 'parts' array in response")
        
        parts = data["parts"]
        
        # MUST have at least 1 entry
        if len(parts) < 1:
            return print_result(False, f"Parts array is empty: {data}")
        
        # Check first part
        first_part = parts[0]
        print(f"\nFirst part details:")
        print(f"  part_no: {first_part.get('part_no')}")
        print(f"  description: {first_part.get('description')}")
        print(f"  mrp: {first_part.get('mrp')}")
        print(f"  image_url: {first_part.get('image_url')}")
        
        # Verify part_no
        if first_part.get("part_no") != "23121KST901S":
            return print_result(False, f"Expected part_no '23121KST901S', got '{first_part.get('part_no')}'")
        
        # Verify description contains GEAR and PRIMARY DRIVE
        description = first_part.get("description", "")
        if "GEAR" not in description.upper() or "PRIMARY DRIVE" not in description.upper():
            return print_result(False, f"Description should contain 'GEAR' and 'PRIMARY DRIVE', got: {description}")
        
        # Verify MRP > 0
        mrp = first_part.get("mrp", 0)
        if mrp <= 0:
            return print_result(False, f"MRP should be > 0, got: {mrp}")
        
        # Verify image_url starts with correct domain
        image_url = first_part.get("image_url", "")
        if not image_url.startswith("https://ecatalogue.heromotocorp.biz"):
            return print_result(False, f"Image URL should start with 'https://ecatalogue.heromotocorp.biz', got: {image_url}")
        
        return print_result(True, f"Hero search successful! Found part {first_part.get('part_no')} with MRP {mrp}")
        
    except Exception as e:
        return print_result(False, f"Exception during Hero search: {str(e)}")

def test_4_hero_search_secondary():
    """Test 4: Hero search with another part - GET /api/hero/search?q=23100KRE900"""
    global owner_token
    
    print_test("4. Hero Search Secondary Part - GET /api/hero/search?q=23100KRE900")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        response = requests.get(
            f"{BASE_URL}/hero/search?q=23100KRE900",
            headers=headers,
            timeout=30
        )
        
        print_response_details(response)
        
        # Must NOT be a 502 ConnectionError
        if response.status_code == 502:
            return print_result(False, "Got 502 error - Hero eCatalogue still unreachable")
        
        if response.status_code != 200:
            return print_result(False, f"Hero search failed with status {response.status_code}: {response.text}")
        
        data = response.json()
        
        # Check parts array exists
        if "parts" not in data:
            return print_result(False, "Missing 'parts' array in response")
        
        parts = data["parts"]
        
        # Empty parts array is OK if part doesn't exist, but must NOT be a connection error
        print(f"Found {len(parts)} parts for 23100KRE900")
        
        if len(parts) > 0:
            print(f"First part: {parts[0]}")
        
        return print_result(True, f"Hero search successful (returned {len(parts)} parts, no connection error)")
        
    except Exception as e:
        return print_result(False, f"Exception during Hero search: {str(e)}")

def test_5_tvs_search_no_regression():
    """Test 5: TVS search still works - GET /api/tvs/search?q=N3012050"""
    global owner_token
    
    print_test("5. TVS Search No Regression - GET /api/tvs/search?q=N3012050")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        response = requests.get(
            f"{BASE_URL}/tvs/search?q=N3012050",
            headers=headers,
            timeout=30
        )
        
        print_response_details(response)
        
        if response.status_code != 200:
            return print_result(False, f"TVS search failed with status {response.status_code}: {response.text}")
        
        data = response.json()
        
        # Check parts array exists
        if "parts" not in data:
            return print_result(False, "Missing 'parts' array in response")
        
        parts = data["parts"]
        
        if len(parts) < 1:
            return print_result(False, f"TVS search returned no parts: {data}")
        
        print(f"TVS search returned {len(parts)} parts")
        print(f"First part: {parts[0]}")
        
        return print_result(True, "TVS search still working correctly (no regression)")
        
    except Exception as e:
        return print_result(False, f"Exception during TVS search: {str(e)}")

def test_6_check_backend_logs():
    """Test 6: Check backend logs for NEW errors after restart"""
    print_test("6. Check Backend Logs for NEW Errors (After Restart)")
    
    try:
        import subprocess
        
        # Get logs after the last "Application startup complete" (which marks the restart)
        result = subprocess.run(
            ["bash", "-c", "tail -n 500 /var/log/supervisor/backend.err.log | tac | sed -n '/Application startup complete/q;p' | tac"],
            capture_output=True,
            text=True,
            timeout=10
        )
        
        log_content_after_restart = result.stdout
        
        print(f"Backend logs AFTER last restart:")
        if log_content_after_restart.strip():
            print(log_content_after_restart)
        else:
            print("(No error logs after restart)")
        
        # Check for specific error patterns in logs AFTER restart
        error_patterns = [
            "Hero eCatalogue unreachable",
            "Name or service not known",
            "Failed to resolve",
            "ecatalogue.heromotocorp.com",
            "ConnectionError",
            "502 Bad Gateway"
        ]
        
        found_errors = []
        for pattern in error_patterns:
            if pattern in log_content_after_restart:
                found_errors.append(pattern)
        
        if found_errors:
            return print_result(False, f"Found NEW error patterns in logs after restart: {found_errors}")
        
        return print_result(True, "No new 502 or DNS resolution errors since backend restart - fix is working!")
        
    except Exception as e:
        return print_result(False, f"Exception checking logs: {str(e)}")

def main():
    """Run all tests"""
    print("\n" + "="*80)
    print("HERO ECATALOGUE URL MIGRATION BUG FIX VERIFICATION")
    print("="*80)
    print(f"Backend URL: {BASE_URL}")
    print(f"Owner credentials: {OWNER_USERNAME}/{OWNER_PASSWORD}")
    print("="*80)
    
    results = []
    
    # Test 1: Login
    results.append(("Login", test_1_login()))
    if not results[-1][1]:
        print("\n❌ Login failed, cannot continue with other tests")
        sys.exit(1)
    
    # Test 2: Check inventory freshness (and upload if needed)
    results.append(("Inventory Freshness", test_2_check_inventory_freshness()))
    
    # Test 3: Hero search happy path
    results.append(("Hero Search Primary (23121KST901)", test_3_hero_search_primary()))
    
    # Test 4: Hero search secondary part
    results.append(("Hero Search Secondary (23100KRE900)", test_4_hero_search_secondary()))
    
    # Test 5: TVS search no regression
    results.append(("TVS Search No Regression", test_5_tvs_search_no_regression()))
    
    # Test 6: Check backend logs
    results.append(("Backend Logs Check", test_6_check_backend_logs()))
    
    # Summary
    print("\n" + "="*80)
    print("TEST SUMMARY")
    print("="*80)
    
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for name, result in results:
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"{status}: {name}")
    
    print("="*80)
    print(f"TOTAL: {passed}/{total} tests passed ({passed*100//total}%)")
    print("="*80)
    
    if passed == total:
        print("\n✅ ALL TESTS PASSED - Hero eCatalogue URL migration bug fix is WORKING")
        sys.exit(0)
    else:
        print(f"\n❌ {total - passed} TEST(S) FAILED - Hero eCatalogue URL migration bug fix still has issues")
        sys.exit(1)

if __name__ == "__main__":
    main()
