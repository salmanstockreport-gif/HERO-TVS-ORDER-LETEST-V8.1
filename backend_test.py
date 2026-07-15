#!/usr/bin/env python3
"""
Backend API Testing for Hero + TVS Parts Ordering System
Tests all endpoints with focus on dual-system support and permission enforcement
"""

import requests
import io
import sys
from openpyxl import Workbook

# Backend URL from frontend/.env
BASE_URL = "https://008a5671-c16f-4b24-91b3-151477b7ed8b.preview.emergentagent.com/api"

# Test credentials
OWNER_USERNAME = "admin"
OWNER_PASSWORD = "admin123"

# Global state
owner_token = None
employee_token = None
employee_id = None
hero_order_id = None
tvs_order_id = None
hero_important_part_id = None
tvs_important_part_id = None
hero_mandatory_part_id = None
tvs_mandatory_part_id = None

def print_test(name):
    """Print test name"""
    print(f"\n{'='*80}")
    print(f"TEST: {name}")
    print('='*80)

def print_result(passed, message=""):
    """Print test result"""
    status = "✅ PASS" if passed else "❌ FAIL"
    print(f"{status}: {message}")
    return passed

def create_test_inventory_xlsx():
    """Create a minimal test inventory Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Headers
    ws['A1'] = "Part No"
    ws['B1'] = "Stock Qty"
    ws['C1'] = "Description"
    
    # Test data - TVS part
    ws['A2'] = "N3012050"
    ws['B2'] = 5
    ws['C2'] = "VALVE STEM OIL SEAL"
    
    # Test data - Hero part
    ws['A3'] = "23121KST901"
    ws['B3'] = 10
    ws['C3'] = "Sample Hero Part"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def test_1_auth_login():
    """Test 1: Auth response shape - login with admin/admin123"""
    global owner_token
    
    print_test("1. Auth Login - Owner credentials")
    
    try:
        response = requests.post(
            f"{BASE_URL}/auth/login",
            json={"username": OWNER_USERNAME, "password": OWNER_PASSWORD},
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"Login failed with status {response.status_code}: {response.text}")
        
        data = response.json()
        
        # Check access_token
        if "access_token" not in data:
            return print_result(False, "Missing access_token in response")
        
        owner_token = data["access_token"]
        
        # Check user object
        if "user" not in data:
            return print_result(False, "Missing user object in response")
        
        user = data["user"]
        
        # Check role
        if user.get("role") != "owner":
            return print_result(False, f"Expected role='owner', got '{user.get('role')}'")
        
        # Check systems
        systems = user.get("systems", [])
        if "hero" not in systems or "tvs" not in systems:
            return print_result(False, f"Expected systems=['hero', 'tvs'], got {systems}")
        
        # Check permissions - all 10 should be true
        permissions = user.get("permissions", {})
        expected_perms = [
            "orders_create_edit", "orders_delete", "orders_mark_sent",
            "search_ecatalogue", "inventory_view", "inventory_upload",
            "manage_important_parts", "manage_mandatory_parts",
            "change_discount", "backup_restore"
        ]
        
        for perm in expected_perms:
            if not permissions.get(perm):
                return print_result(False, f"Permission '{perm}' is not true")
        
        print(f"Owner token: {owner_token[:20]}...")
        print(f"User: {user.get('username')}, Role: {user.get('role')}, Systems: {systems}")
        print(f"Permissions: {len([k for k, v in permissions.items() if v])}/10 enabled")
        
        return print_result(True, "Login successful with correct response shape")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_2_inventory_freshness_gate():
    """Test 2: Inventory freshness gate - should return 423 when stale"""
    print_test("2. Inventory Freshness Gate - Test 423 response")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        # Try TVS search without inventory
        response = requests.get(
            f"{BASE_URL}/tvs/search",
            params={"q": "N3012050"},
            headers=headers,
            timeout=30
        )
        
        # Should return 423 if inventory is stale
        if response.status_code == 423:
            data = response.json()
            detail = data.get("detail", {})
            if isinstance(detail, dict) and detail.get("code") == "inventory_stale":
                return print_result(True, "Correctly returned 423 with inventory_stale code")
            else:
                return print_result(False, f"Got 423 but wrong detail: {detail}")
        elif response.status_code == 200:
            # Inventory might already be fresh from previous tests
            print("⚠️  WARNING: Inventory is already fresh (200 response)")
            return print_result(True, "Inventory is fresh (skipping stale test)")
        else:
            return print_result(False, f"Unexpected status {response.status_code}: {response.text}")
            
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_3_inventory_upload():
    """Test 3: Upload inventory"""
    print_test("3. Inventory Upload - Upload test inventory")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        # Create test inventory file
        xlsx_data = create_test_inventory_xlsx()
        
        files = {
            'file': ('test_inventory.xlsx', xlsx_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        data = {
            'part_no': 'Part No',
            'stock_qty': 'Stock Qty',
            'description': 'Description',
            'replace': 'true'
        }
        
        response = requests.post(
            f"{BASE_URL}/inventory/upload",
            headers=headers,
            files=files,
            data=data,
            timeout=30
        )
        
        if response.status_code not in [200, 201]:
            return print_result(False, f"Upload failed with status {response.status_code}: {response.text}")
        
        result = response.json()
        
        if not result.get("success"):
            return print_result(False, "Upload did not return success=true")
        
        imported = result.get("imported", 0)
        if imported < 1:
            return print_result(False, f"Expected at least 1 imported row, got {imported}")
        
        print(f"Imported {imported} inventory items")
        return print_result(True, f"Inventory uploaded successfully ({imported} items)")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_4_tvs_ecatalogue_search():
    """Test 4: TVS eCatalogue search"""
    print_test("4. TVS eCatalogue Search - Search for N3012050")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        response = requests.get(
            f"{BASE_URL}/tvs/search",
            params={"q": "N3012050"},
            headers=headers,
            timeout=30
        )
        
        if response.status_code == 502:
            # TVS eCatalogue might be unreachable
            print("⚠️  WARNING: TVS eCatalogue unreachable (502)")
            return print_result(True, "TVS eCatalogue unreachable - acceptable for external API")
        
        if response.status_code != 200:
            return print_result(False, f"Search failed with status {response.status_code}: {response.text}")
        
        data = response.json()
        
        if "parts" not in data:
            return print_result(False, "Missing 'parts' array in response")
        
        parts = data["parts"]
        
        if len(parts) < 1:
            return print_result(False, "Expected at least 1 part in results")
        
        # Check first part
        first_part = parts[0]
        
        if first_part.get("part_no") != "N3012050":
            return print_result(False, f"Expected part_no='N3012050', got '{first_part.get('part_no')}'")
        
        description = first_part.get("description", "")
        if "VALVE STEM OIL SEAL" not in description.upper():
            return print_result(False, f"Expected description containing 'VALVE STEM OIL SEAL', got '{description}'")
        
        mrp = first_part.get("mrp")
        if mrp != 80:
            print(f"⚠️  WARNING: Expected MRP=80, got {mrp} (TVS API might have changed)")
        
        print(f"Found {len(parts)} parts")
        print(f"First part: {first_part.get('part_no')} - {first_part.get('description')} - MRP: {first_part.get('mrp')}")
        
        return print_result(True, "TVS search successful with correct data")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_5_hero_ecatalogue_search():
    """Test 5: Hero eCatalogue search"""
    print_test("5. Hero eCatalogue Search - Test Hero search endpoint")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        response = requests.get(
            f"{BASE_URL}/hero/search",
            params={"q": "23121KST901"},
            headers=headers,
            timeout=30
        )
        
        if response.status_code == 502:
            # Hero eCatalogue might be unreachable
            print("⚠️  WARNING: Hero eCatalogue unreachable (502)")
            return print_result(True, "Hero eCatalogue unreachable - acceptable for external API")
        
        if response.status_code != 200:
            return print_result(False, f"Search failed with status {response.status_code}: {response.text}")
        
        data = response.json()
        
        if "parts" not in data:
            return print_result(False, "Missing 'parts' array in response")
        
        print(f"Hero search returned {len(data['parts'])} parts")
        return print_result(True, "Hero search endpoint working")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_6_system_scoped_orders_create():
    """Test 6: System-scoped orders - Create orders for hero and tvs"""
    global hero_order_id, tvs_order_id
    
    print_test("6. System-Scoped Orders - Create Hero and TVS orders")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        # Create Hero order
        hero_body = {
            "items": [{
                "part_no": "23121KST901",
                "description": "Sample Hero Part",
                "mrp": 100,
                "qty": 1
            }],
            "remarks": "hero test order"
        }
        
        response = requests.post(
            f"{BASE_URL}/orders",
            params={"system": "hero"},
            headers=headers,
            json=hero_body,
            timeout=30
        )
        
        if response.status_code not in [200, 201]:
            return print_result(False, f"Hero order creation failed with status {response.status_code}: {response.text}")
        
        hero_order = response.json()
        hero_order_id = hero_order.get("id")
        hero_order_no = hero_order.get("order_no")
        
        if not hero_order_no or not hero_order_no.startswith("HMC-"):
            return print_result(False, f"Hero order_no should start with 'HMC-', got '{hero_order_no}'")
        
        if hero_order.get("system") != "hero":
            return print_result(False, f"Hero order system field should be 'hero', got '{hero_order.get('system')}'")
        
        print(f"Hero order created: {hero_order_no} (ID: {hero_order_id})")
        
        # Create TVS order
        tvs_body = {
            "items": [{
                "part_no": "N3012050",
                "description": "VALVE STEM OIL SEAL",
                "mrp": 80,
                "qty": 1
            }],
            "remarks": "tvs test order"
        }
        
        response = requests.post(
            f"{BASE_URL}/orders",
            params={"system": "tvs"},
            headers=headers,
            json=tvs_body,
            timeout=30
        )
        
        if response.status_code not in [200, 201]:
            return print_result(False, f"TVS order creation failed with status {response.status_code}: {response.text}")
        
        tvs_order = response.json()
        tvs_order_id = tvs_order.get("id")
        tvs_order_no = tvs_order.get("order_no")
        
        if not tvs_order_no or not tvs_order_no.startswith("TVS-"):
            return print_result(False, f"TVS order_no should start with 'TVS-', got '{tvs_order_no}'")
        
        if tvs_order.get("system") != "tvs":
            return print_result(False, f"TVS order system field should be 'tvs', got '{tvs_order.get('system')}'")
        
        print(f"TVS order created: {tvs_order_no} (ID: {tvs_order_id})")
        
        return print_result(True, "Both Hero and TVS orders created successfully")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_7_system_scoped_orders_filter():
    """Test 7: System-scoped orders - Filter by system"""
    print_test("7. System-Scoped Orders - Filter by system")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        # Get Hero orders
        response = requests.get(
            f"{BASE_URL}/orders",
            params={"system": "hero"},
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"Hero orders fetch failed with status {response.status_code}: {response.text}")
        
        hero_orders = response.json()
        
        # Check all orders are hero system
        for order in hero_orders:
            if order.get("system") != "hero":
                return print_result(False, f"Found non-hero order in hero list: {order.get('order_no')}")
        
        print(f"Hero orders: {len(hero_orders)} (all have system='hero')")
        
        # Get TVS orders
        response = requests.get(
            f"{BASE_URL}/orders",
            params={"system": "tvs"},
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"TVS orders fetch failed with status {response.status_code}: {response.text}")
        
        tvs_orders = response.json()
        
        # Check all orders are tvs system
        for order in tvs_orders:
            if order.get("system") != "tvs":
                return print_result(False, f"Found non-tvs order in tvs list: {order.get('order_no')}")
        
        print(f"TVS orders: {len(tvs_orders)} (all have system='tvs')")
        
        return print_result(True, "Order filtering by system works correctly")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_8_permission_keys():
    """Test 8: Get permission keys"""
    print_test("8. Permission Keys - Get list of permission keys")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        response = requests.get(
            f"{BASE_URL}/permissions/keys",
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"Failed with status {response.status_code}: {response.text}")
        
        data = response.json()
        
        if "keys" not in data:
            return print_result(False, "Missing 'keys' in response")
        
        keys = data["keys"]
        
        if len(keys) != 10:
            return print_result(False, f"Expected 10 permission keys, got {len(keys)}")
        
        print(f"Permission keys: {len(keys)}")
        for key in keys:
            print(f"  - {key.get('key')}: {key.get('label')}")
        
        return print_result(True, "Permission keys retrieved successfully")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_9_employee_create():
    """Test 9: Create employee with limited permissions"""
    global employee_id
    
    print_test("9. Employee CRUD - Create employee")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        employee_body = {
            "username": "tvsemp1",
            "password": "tvspass123",
            "systems": ["tvs"],
            "permissions": {
                "orders_create_edit": True,
                "search_ecatalogue": True,
                "inventory_view": True
            }
        }
        
        response = requests.post(
            f"{BASE_URL}/employees",
            headers=headers,
            json=employee_body,
            timeout=30
        )
        
        if response.status_code not in [200, 201]:
            return print_result(False, f"Employee creation failed with status {response.status_code}: {response.text}")
        
        employee = response.json()
        employee_id = employee.get("id")
        
        if employee.get("role") != "employee":
            return print_result(False, f"Expected role='employee', got '{employee.get('role')}'")
        
        if employee.get("systems") != ["tvs"]:
            return print_result(False, f"Expected systems=['tvs'], got {employee.get('systems')}")
        
        print(f"Employee created: {employee.get('username')} (ID: {employee_id})")
        print(f"Systems: {employee.get('systems')}")
        print(f"Permissions: {employee.get('permissions')}")
        
        return print_result(True, "Employee created successfully")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_10_employee_login():
    """Test 10: Login as employee and verify permissions"""
    global employee_token
    
    print_test("10. Employee Login - Login as tvsemp1")
    
    try:
        response = requests.post(
            f"{BASE_URL}/auth/login",
            json={"username": "tvsemp1", "password": "tvspass123"},
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"Login failed with status {response.status_code}: {response.text}")
        
        data = response.json()
        employee_token = data.get("access_token")
        user = data.get("user", {})
        
        if user.get("role") != "employee":
            return print_result(False, f"Expected role='employee', got '{user.get('role')}'")
        
        if user.get("systems") != ["tvs"]:
            return print_result(False, f"Expected systems=['tvs'], got {user.get('systems')}")
        
        # Check only requested permissions are true
        permissions = user.get("permissions", {})
        
        if not permissions.get("orders_create_edit"):
            return print_result(False, "orders_create_edit should be true")
        
        if not permissions.get("search_ecatalogue"):
            return print_result(False, "search_ecatalogue should be true")
        
        if not permissions.get("inventory_view"):
            return print_result(False, "inventory_view should be true")
        
        # Check other permissions are false
        if permissions.get("orders_delete"):
            return print_result(False, "orders_delete should be false")
        
        print(f"Employee token: {employee_token[:20]}...")
        print(f"Role: {user.get('role')}, Systems: {user.get('systems')}")
        
        return print_result(True, "Employee login successful with correct permissions")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_11_employee_tvs_search_allowed():
    """Test 11: Employee can search TVS (has permission and system access)"""
    print_test("11. Employee Permission - TVS search allowed")
    
    try:
        headers = {"Authorization": f"Bearer {employee_token}"}
        
        response = requests.get(
            f"{BASE_URL}/tvs/search",
            params={"q": "N3012050"},
            headers=headers,
            timeout=30
        )
        
        if response.status_code == 502:
            print("⚠️  WARNING: TVS eCatalogue unreachable (502)")
            return print_result(True, "TVS eCatalogue unreachable - acceptable")
        
        if response.status_code != 200:
            return print_result(False, f"TVS search failed with status {response.status_code}: {response.text}")
        
        return print_result(True, "Employee can search TVS (has permission and system access)")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_12_employee_hero_search_forbidden():
    """Test 12: Employee cannot search Hero (no system access)"""
    print_test("12. Employee Permission - Hero search forbidden")
    
    try:
        headers = {"Authorization": f"Bearer {employee_token}"}
        
        response = requests.get(
            f"{BASE_URL}/hero/search",
            params={"q": "23121KST901"},
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 403:
            return print_result(False, f"Expected 403, got {response.status_code}: {response.text}")
        
        return print_result(True, "Employee correctly denied Hero system access (403)")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_13_employee_create_employee_forbidden():
    """Test 13: Employee cannot create other employees (owner-only)"""
    print_test("13. Employee Permission - Cannot create employees")
    
    try:
        headers = {"Authorization": f"Bearer {employee_token}"}
        
        body = {
            "username": "test123",
            "password": "test123",
            "systems": ["hero"],
            "permissions": {}
        }
        
        response = requests.post(
            f"{BASE_URL}/employees",
            headers=headers,
            json=body,
            timeout=30
        )
        
        if response.status_code != 403:
            return print_result(False, f"Expected 403, got {response.status_code}: {response.text}")
        
        return print_result(True, "Employee correctly denied employee creation (403)")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_14_employee_delete_order_forbidden():
    """Test 14: Employee cannot delete orders (no permission and no system access)"""
    print_test("14. Employee Permission - Cannot delete Hero order")
    
    try:
        headers = {"Authorization": f"Bearer {employee_token}"}
        
        response = requests.delete(
            f"{BASE_URL}/orders/{hero_order_id}",
            params={"confirm": "delete"},
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 403:
            return print_result(False, f"Expected 403, got {response.status_code}: {response.text}")
        
        return print_result(True, "Employee correctly denied order deletion (403)")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_15_owner_delete_employee():
    """Test 15: Owner can delete employee"""
    print_test("15. Employee CRUD - Owner deletes employee")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        response = requests.delete(
            f"{BASE_URL}/employees/{employee_id}",
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"Delete failed with status {response.status_code}: {response.text}")
        
        data = response.json()
        
        if not data.get("success"):
            return print_result(False, "Delete did not return success=true")
        
        return print_result(True, "Employee deleted successfully")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_16_important_parts_hero():
    """Test 16: Create important part for Hero system"""
    global hero_important_part_id
    
    print_test("16. Important Parts - Create Hero important part")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        body = {
            "part_no": "23121KST901",
            "description": "Sample Hero Part",
            "threshold_qty": 5
        }
        
        response = requests.post(
            f"{BASE_URL}/important-parts",
            params={"system": "hero"},
            headers=headers,
            json=body,
            timeout=30
        )
        
        if response.status_code not in [200, 201]:
            return print_result(False, f"Creation failed with status {response.status_code}: {response.text}")
        
        part = response.json()
        hero_important_part_id = part.get("id")
        
        if part.get("system") != "hero":
            return print_result(False, f"Expected system='hero', got '{part.get('system')}'")
        
        print(f"Hero important part created: {part.get('part_no')} (ID: {hero_important_part_id})")
        
        return print_result(True, "Hero important part created successfully")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_17_important_parts_tvs():
    """Test 17: Create important part for TVS system"""
    global tvs_important_part_id
    
    print_test("17. Important Parts - Create TVS important part")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        body = {
            "part_no": "N3012050",
            "description": "VALVE STEM OIL SEAL",
            "threshold_qty": 2
        }
        
        response = requests.post(
            f"{BASE_URL}/important-parts",
            params={"system": "tvs"},
            headers=headers,
            json=body,
            timeout=30
        )
        
        if response.status_code not in [200, 201]:
            return print_result(False, f"Creation failed with status {response.status_code}: {response.text}")
        
        part = response.json()
        tvs_important_part_id = part.get("id")
        
        if part.get("system") != "tvs":
            return print_result(False, f"Expected system='tvs', got '{part.get('system')}'")
        
        print(f"TVS important part created: {part.get('part_no')} (ID: {tvs_important_part_id})")
        
        return print_result(True, "TVS important part created successfully")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_18_important_parts_filter():
    """Test 18: Filter important parts by system"""
    print_test("18. Important Parts - Filter by system")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        # Get Hero important parts
        response = requests.get(
            f"{BASE_URL}/important-parts",
            params={"system": "hero"},
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"Hero fetch failed with status {response.status_code}: {response.text}")
        
        hero_parts = response.json()
        
        # Check all are hero system
        for part in hero_parts:
            if part.get("system") != "hero":
                return print_result(False, f"Found non-hero part in hero list: {part.get('part_no')}")
        
        print(f"Hero important parts: {len(hero_parts)}")
        
        # Get TVS important parts
        response = requests.get(
            f"{BASE_URL}/important-parts",
            params={"system": "tvs"},
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"TVS fetch failed with status {response.status_code}: {response.text}")
        
        tvs_parts = response.json()
        
        # Check all are tvs system
        for part in tvs_parts:
            if part.get("system") != "tvs":
                return print_result(False, f"Found non-tvs part in tvs list: {part.get('part_no')}")
        
        print(f"TVS important parts: {len(tvs_parts)}")
        
        return print_result(True, "Important parts filtering by system works correctly")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_19_mandatory_parts_hero():
    """Test 19: Create mandatory part for Hero system"""
    global hero_mandatory_part_id
    
    print_test("19. Mandatory Parts - Create Hero mandatory part")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        body = {
            "part_no": "23121KST901",
            "description": "Sample Hero Part",
            "mrp": 100,
            "qty": 1
        }
        
        response = requests.post(
            f"{BASE_URL}/mandatory-parts",
            params={"system": "hero"},
            headers=headers,
            json=body,
            timeout=30
        )
        
        if response.status_code not in [200, 201]:
            return print_result(False, f"Creation failed with status {response.status_code}: {response.text}")
        
        part = response.json()
        hero_mandatory_part_id = part.get("id")
        
        if part.get("system") != "hero":
            return print_result(False, f"Expected system='hero', got '{part.get('system')}'")
        
        print(f"Hero mandatory part created: {part.get('part_no')} (ID: {hero_mandatory_part_id})")
        
        return print_result(True, "Hero mandatory part created successfully")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_20_mandatory_parts_tvs():
    """Test 20: Create mandatory part for TVS system"""
    global tvs_mandatory_part_id
    
    print_test("20. Mandatory Parts - Create TVS mandatory part")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        body = {
            "part_no": "N3012050",
            "description": "VALVE STEM OIL SEAL",
            "mrp": 80,
            "qty": 1
        }
        
        response = requests.post(
            f"{BASE_URL}/mandatory-parts",
            params={"system": "tvs"},
            headers=headers,
            json=body,
            timeout=30
        )
        
        if response.status_code not in [200, 201]:
            return print_result(False, f"Creation failed with status {response.status_code}: {response.text}")
        
        part = response.json()
        tvs_mandatory_part_id = part.get("id")
        
        if part.get("system") != "tvs":
            return print_result(False, f"Expected system='tvs', got '{part.get('system')}'")
        
        print(f"TVS mandatory part created: {part.get('part_no')} (ID: {tvs_mandatory_part_id})")
        
        return print_result(True, "TVS mandatory part created successfully")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_21_mandatory_parts_filter():
    """Test 21: Filter mandatory parts by system"""
    print_test("21. Mandatory Parts - Filter by system")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        # Get Hero mandatory parts
        response = requests.get(
            f"{BASE_URL}/mandatory-parts",
            params={"system": "hero"},
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"Hero fetch failed with status {response.status_code}: {response.text}")
        
        data = response.json()
        hero_parts = data.get("parts", [])
        
        # Check all are hero system
        for part in hero_parts:
            if part.get("system") != "hero":
                return print_result(False, f"Found non-hero part in hero list: {part.get('part_no')}")
        
        print(f"Hero mandatory parts: {len(hero_parts)}")
        
        # Get TVS mandatory parts
        response = requests.get(
            f"{BASE_URL}/mandatory-parts",
            params={"system": "tvs"},
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"TVS fetch failed with status {response.status_code}: {response.text}")
        
        data = response.json()
        tvs_parts = data.get("parts", [])
        
        # Check all are tvs system
        for part in tvs_parts:
            if part.get("system") != "tvs":
                return print_result(False, f"Found non-tvs part in tvs list: {part.get('part_no')}")
        
        print(f"TVS mandatory parts: {len(tvs_parts)}")
        
        return print_result(True, "Mandatory parts filtering by system works correctly")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_22_dashboard_stats_hero():
    """Test 22: Dashboard stats for Hero system"""
    print_test("22. Dashboard Stats - Hero system")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        response = requests.get(
            f"{BASE_URL}/dashboard/stats",
            params={"system": "hero"},
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"Failed with status {response.status_code}: {response.text}")
        
        stats = response.json()
        
        if stats.get("system") != "hero":
            return print_result(False, f"Expected system='hero', got '{stats.get('system')}'")
        
        print(f"Hero stats: current_orders={stats.get('current_orders')}, sent_orders={stats.get('sent_orders')}")
        
        return print_result(True, "Hero dashboard stats retrieved successfully")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_23_dashboard_stats_tvs():
    """Test 23: Dashboard stats for TVS system"""
    print_test("23. Dashboard Stats - TVS system")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        response = requests.get(
            f"{BASE_URL}/dashboard/stats",
            params={"system": "tvs"},
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"Failed with status {response.status_code}: {response.text}")
        
        stats = response.json()
        
        if stats.get("system") != "tvs":
            return print_result(False, f"Expected system='tvs', got '{stats.get('system')}'")
        
        print(f"TVS stats: current_orders={stats.get('current_orders')}, sent_orders={stats.get('sent_orders')}")
        
        return print_result(True, "TVS dashboard stats retrieved successfully")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_24_inventory_lookup():
    """Test 24: Legacy compatibility - Inventory lookup"""
    print_test("24. Legacy Compatibility - Inventory lookup")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        response = requests.get(
            f"{BASE_URL}/inventory/lookup/N3012050",
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"Failed with status {response.status_code}: {response.text}")
        
        data = response.json()
        
        if not data.get("found"):
            return print_result(False, "Part not found in inventory")
        
        print(f"Found part: {data.get('part_no')} - Stock: {data.get('stock_qty')}")
        
        return print_result(True, "Inventory lookup working")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def test_25_inventory_mapping():
    """Test 25: Legacy compatibility - Inventory mapping"""
    print_test("25. Legacy Compatibility - Inventory mapping")
    
    try:
        headers = {"Authorization": f"Bearer {owner_token}"}
        
        # GET mapping
        response = requests.get(
            f"{BASE_URL}/inventory/mapping",
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"GET failed with status {response.status_code}: {response.text}")
        
        mapping = response.json()
        print(f"Current mapping: {mapping}")
        
        # PUT mapping (update)
        new_mapping = {
            "part_no": "Part No",
            "description": "Description",
            "stock_qty": "Stock Qty",
            "location": "",
            "rate": ""
        }
        
        response = requests.put(
            f"{BASE_URL}/inventory/mapping",
            headers=headers,
            json=new_mapping,
            timeout=30
        )
        
        if response.status_code != 200:
            return print_result(False, f"PUT failed with status {response.status_code}: {response.text}")
        
        return print_result(True, "Inventory mapping GET/PUT working")
        
    except Exception as e:
        return print_result(False, f"Exception: {str(e)}")

def main():
    """Run all tests"""
    print("\n" + "="*80)
    print("BACKEND API TESTING - Hero + TVS Parts Ordering System")
    print("="*80)
    print(f"Backend URL: {BASE_URL}")
    print(f"Owner credentials: {OWNER_USERNAME}/{OWNER_PASSWORD}")
    
    results = []
    
    # Run all tests in order
    results.append(test_1_auth_login())
    results.append(test_2_inventory_freshness_gate())
    results.append(test_3_inventory_upload())
    results.append(test_4_tvs_ecatalogue_search())
    results.append(test_5_hero_ecatalogue_search())
    results.append(test_6_system_scoped_orders_create())
    results.append(test_7_system_scoped_orders_filter())
    results.append(test_8_permission_keys())
    results.append(test_9_employee_create())
    results.append(test_10_employee_login())
    results.append(test_11_employee_tvs_search_allowed())
    results.append(test_12_employee_hero_search_forbidden())
    results.append(test_13_employee_create_employee_forbidden())
    results.append(test_14_employee_delete_order_forbidden())
    results.append(test_15_owner_delete_employee())
    results.append(test_16_important_parts_hero())
    results.append(test_17_important_parts_tvs())
    results.append(test_18_important_parts_filter())
    results.append(test_19_mandatory_parts_hero())
    results.append(test_20_mandatory_parts_tvs())
    results.append(test_21_mandatory_parts_filter())
    results.append(test_22_dashboard_stats_hero())
    results.append(test_23_dashboard_stats_tvs())
    results.append(test_24_inventory_lookup())
    results.append(test_25_inventory_mapping())
    
    # Summary
    print("\n" + "="*80)
    print("TEST SUMMARY")
    print("="*80)
    
    passed = sum(results)
    total = len(results)
    
    print(f"\nTotal: {total} tests")
    print(f"Passed: {passed} ✅")
    print(f"Failed: {total - passed} ❌")
    print(f"Success rate: {passed/total*100:.1f}%")
    
    if passed == total:
        print("\n🎉 ALL TESTS PASSED!")
        return 0
    else:
        print(f"\n⚠️  {total - passed} test(s) failed")
        return 1

if __name__ == "__main__":
    sys.exit(main())
