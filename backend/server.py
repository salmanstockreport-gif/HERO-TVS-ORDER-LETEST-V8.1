from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import logging
import uuid
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Dict, Any

import bcrypt
import jwt
import requests
from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage

# ---------------------------------------------------------------------------
# App / DB Setup
# ---------------------------------------------------------------------------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="Hero Parts Ordering System")
api_router = APIRouter(prefix="/api")
security = HTTPBearer(auto_error=False)

JWT_ALGORITHM = "HS256"
JWT_SECRET = os.environ["JWT_SECRET"]
HERO_URL = os.environ["HERO_ECATALOGUE_URL"]
TVS_URL = os.environ.get("TVS_ECOMMERCE_API_URL", "https://www.advantagetvs.com/PartEcommerceAPI/")
TVS_DEALER_ID = int(os.environ.get("TVS_DEALER_ID", "10001"))
TVS_BRANCH_ID = int(os.environ.get("TVS_BRANCH_ID", "1"))
TVS_CUSTOMER_TYPE = os.environ.get("TVS_CUSTOMER_TYPE", "Customer")

# Supported systems
SYSTEMS = ("hero", "tvs")

# Employee permission keys. Owner accounts always bypass these checks.
PERMISSION_KEYS = [
    "orders_create_edit",
    "orders_delete",
    "orders_mark_sent",
    "search_ecatalogue",
    "inventory_view",
    "inventory_upload",
    "manage_important_parts",
    "manage_mandatory_parts",
    "change_discount",
    "backup_restore",
]


def default_permissions(all_true: bool = False) -> dict:
    return {k: bool(all_true) for k in PERMISSION_KEYS}

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(pw: str, hashed: str) -> bool:
    return bcrypt.checkpw(pw.encode("utf-8"), hashed.encode("utf-8"))


def create_access_token(user_id: str, username: str) -> str:
    payload = {
        "sub": user_id,
        "username": username,
        "exp": datetime.now(timezone.utc) + timedelta(hours=12),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_user(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> dict:
    if credentials is None:
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = credentials.credentials
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    # Legacy users may not have role/systems/permissions -- default to owner (backwards compat)
    user.setdefault("role", "owner")
    user.setdefault("systems", list(SYSTEMS))
    perms = default_permissions(all_true=(user["role"] == "owner"))
    perms.update(user.get("permissions") or {})
    user["permissions"] = perms
    return user


def is_owner(user: dict) -> bool:
    return user.get("role") == "owner"


def require_owner(current_user: dict = Depends(get_current_user)) -> dict:
    if not is_owner(current_user):
        raise HTTPException(status_code=403, detail="Owner access required")
    return current_user


def require_permission(perm: str):
    """Return a dependency that ensures the current user has the given permission
    (owners bypass). Also returns the user dict."""

    async def _dep(current_user: dict = Depends(get_current_user)) -> dict:
        if is_owner(current_user):
            return current_user
        if not current_user.get("permissions", {}).get(perm):
            raise HTTPException(
                status_code=403,
                detail=f"Missing permission: {perm}",
            )
        return current_user

    return _dep


def require_system_access(user: dict, system: str) -> None:
    if system not in SYSTEMS:
        raise HTTPException(status_code=400, detail=f"Invalid system: {system}")
    if is_owner(user):
        return
    if system not in (user.get("systems") or []):
        raise HTTPException(status_code=403, detail=f"No access to {system} system")


# ---------------------------------------------------------------------------
# Hero eCatalogue client
# ---------------------------------------------------------------------------
class HeroClient:
    def __init__(self, base_url: str):
        self.base_url = base_url.rstrip("/") + "/"
        self._token: Optional[str] = None
        self._user_code: Optional[str] = None
        self._token_ts: Optional[datetime] = None

    def _refresh_token(self) -> None:
        r = requests.post(
            self.base_url + "getGeneralUserDetails",
            headers={"Content-Type": "application/json"},
            timeout=20,
        )
        r.raise_for_status()
        data = r.json()
        self._token = data.get("authToken")
        self._user_code = data.get("userId") or "General"
        self._token_ts = datetime.now(timezone.utc)

    def _ensure_token(self) -> None:
        if self._token is None or self._token_ts is None:
            self._refresh_token()
            return
        # refresh every 6 hours to be safe
        if datetime.now(timezone.utc) - self._token_ts > timedelta(hours=6):
            self._refresh_token()

    def search_part(self, part_no: str) -> Dict[str, Any]:
        self._ensure_token()
        payload = {
            "userCode": self._user_code,
            "partno": part_no,
            "supercedence": "",
            "language": "1",
            "headerSearch": True,
        }
        headers = {
            "Content-Type": "application/json",
            "X-Auth-Token": self._token,
        }
        r = requests.post(self.base_url + "partSearch", json=payload, headers=headers, timeout=25)
        if r.status_code == 401 or r.status_code == 403:
            self._refresh_token()
            headers["X-Auth-Token"] = self._token
            r = requests.post(self.base_url + "partSearch", json=payload, headers=headers, timeout=25)
        r.raise_for_status()
        return r.json()


hero_client = HeroClient(HERO_URL)


# ---------------------------------------------------------------------------
# TVS eCatalogue client (advantagetvs.com PartEcommerceAPI)
# ---------------------------------------------------------------------------
class TVSClient:
    def __init__(self, base_url: str, dealer_id: int, branch_id: int, cust_type: str):
        self.base_url = base_url.rstrip("/") + "/"
        self.dealer_id = dealer_id
        self.branch_id = branch_id
        self.cust_type = cust_type
        self._token: Optional[str] = None
        self._token_ts: Optional[datetime] = None

    def _refresh_token(self) -> None:
        r = requests.post(
            self.base_url + "Setting/tokenGeneration",
            headers={"Content-Type": "application/json"},
            json={
                "dealerId": self.dealer_id,
                "branchId": self.branch_id,
                "Type": self.cust_type,
            },
            timeout=20,
        )
        r.raise_for_status()
        data = r.json()
        self._token = data.get("access_token")
        self._token_ts = datetime.now(timezone.utc)

    def _ensure_token(self) -> None:
        if self._token is None or self._token_ts is None:
            self._refresh_token()
            return
        # Access tokens refresh every 30 minutes (TVS runtime uses similar cadence)
        if datetime.now(timezone.utc) - self._token_ts > timedelta(minutes=25):
            self._refresh_token()

    def search_part(self, part_no: str) -> Dict[str, Any]:
        self._ensure_token()
        params = {
            "partid": (part_no or "").strip(),
            "description": "",
            "partdesc": "",
            "partSeries": "",
            "modelID": "",
            "page": 1,
            "pageSize": 100,
            "frameNumber": "",
        }
        headers = {
            "Content-Type": "text/plain",
            "Authorization": f"Bearer {self._token}",
        }
        r = requests.get(
            self.base_url + "api/Catalouge/GetPartsearch",
            params=params,
            headers=headers,
            timeout=25,
        )
        if r.status_code in (401, 403):
            self._refresh_token()
            headers["Authorization"] = f"Bearer {self._token}"
            r = requests.get(
                self.base_url + "api/Catalouge/GetPartsearch",
                params=params,
                headers=headers,
                timeout=25,
            )
        r.raise_for_status()
        return r.json()


tvs_client = TVSClient(TVS_URL, TVS_DEALER_ID, TVS_BRANCH_ID, TVS_CUSTOMER_TYPE)


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class LoginRequest(BaseModel):
    username: str
    password: str


class CredentialsUpdate(BaseModel):
    current_password: str
    new_username: Optional[str] = None
    new_password: Optional[str] = None


class DiscountUpdate(BaseModel):
    discount_percent: float


class OrderItem(BaseModel):
    part_no: str
    description: str = ""
    mrp: float = 0.0
    qty: int = 1
    discount_percent: Optional[float] = None
    landed_price: float = 0.0
    line_total: float = 0.0
    moq: Optional[float] = None
    note: str = ""


class OrderCreate(BaseModel):
    items: List[OrderItem] = []
    remarks: str = ""


class OrderUpdate(BaseModel):
    items: List[OrderItem]
    remarks: str = ""


class InventoryMapping(BaseModel):
    part_no: str = "Part No"
    description: str = "Description"
    stock_qty: str = "Stock Qty"
    location: str = ""
    rate: str = ""


class ImportantPartBody(BaseModel):
    part_no: str
    description: str = ""
    threshold_qty: float = 1


class MandatoryPartBody(BaseModel):
    part_no: str
    description: str = ""
    mrp: float = 0.0
    qty: int = 1


class MandatoryToggleBody(BaseModel):
    enabled: bool


# ------ Employee / user-management payloads ------
class EmployeeCreate(BaseModel):
    username: str
    password: str
    systems: List[str] = ["hero"]
    permissions: Dict[str, bool] = {}


class EmployeeUpdate(BaseModel):
    password: Optional[str] = None
    systems: Optional[List[str]] = None
    permissions: Optional[Dict[str, bool]] = None


# Inventory freshness window (hours)
INVENTORY_TTL_HOURS = 24

# Maximum number of concurrent CURRENT (draft) orders allowed. When the count
# hits this value, POST /orders is refused with HTTP 409 until at least one is
# marked sent or deleted.
MAX_CURRENT_ORDERS = 2


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_part_no(pn: str) -> str:
    return (pn or "").strip().upper().replace(" ", "").replace("-", "")


def format_part_no_display(pn: str) -> str:
    """Customer display format: remove dashes/spaces, append 'S' if missing."""
    if not pn:
        return ""
    p = str(pn).strip().upper().replace(" ", "").replace("-", "")
    if not p.endswith("S"):
        p = p + "S"
    return p


SYSTEM_ORDER_PREFIX = {"hero": "HMC", "tvs": "TVS"}


async def generate_order_no(system: str = "hero") -> str:
    prefix_root = SYSTEM_ORDER_PREFIX.get(system, "HMC")
    date_part = datetime.now(timezone.utc).strftime("%Y%m%d")
    prefix = f"{prefix_root}-{date_part}-"
    # find last order for today for this system
    cursor = db.orders.find({"order_no": {"$regex": f"^{prefix}"}}).sort("order_no", -1).limit(1)
    docs = await cursor.to_list(1)
    if docs:
        last_seq = int(docs[0]["order_no"].split("-")[-1])
        seq = last_seq + 1
    else:
        seq = 1
    return f"{prefix}{seq:03d}"


def compute_item_totals(item: dict, global_discount: float) -> dict:
    mrp = float(item.get("mrp") or 0)
    qty = int(item.get("qty") or 0)
    raw_disc = item.get("discount_percent")
    if raw_disc is None or raw_disc == "":
        disc = float(global_discount)
    else:
        disc = float(raw_disc)
    landed = round(mrp * (1 - disc / 100.0), 2)
    line_total = round(landed * qty, 2)
    item["discount_percent"] = disc
    item["landed_price"] = landed
    item["line_total"] = line_total
    return item


# ---------------------------------------------------------------------------
# Inventory freshness (24h lock)
# ---------------------------------------------------------------------------
async def get_inventory_status() -> dict:
    """Return whether the inventory is fresh (< INVENTORY_TTL_HOURS old)."""
    doc = await db.settings.find_one({"key": "inventory_status"}, {"_id": 0})
    last = doc.get("last_uploaded_at") if doc else None
    if not last:
        return {
            "fresh": False,
            "last_uploaded_at": None,
            "expires_at": None,
            "hours_remaining": 0,
            "ttl_hours": INVENTORY_TTL_HOURS,
            "never_uploaded": True,
        }
    try:
        last_dt = datetime.fromisoformat(last)
        if last_dt.tzinfo is None:
            last_dt = last_dt.replace(tzinfo=timezone.utc)
    except Exception:
        return {
            "fresh": False,
            "last_uploaded_at": last,
            "expires_at": None,
            "hours_remaining": 0,
            "ttl_hours": INVENTORY_TTL_HOURS,
            "never_uploaded": False,
        }
    assert last_dt is not None  # narrows for static analyzers; try/except above guarantees this
    expires = last_dt + timedelta(hours=INVENTORY_TTL_HOURS)
    now = datetime.now(timezone.utc)
    hours_remaining = max(0.0, (expires - now).total_seconds() / 3600.0)
    return {
        "fresh": now < expires,
        "last_uploaded_at": last_dt.isoformat(),
        "expires_at": expires.isoformat(),
        "hours_remaining": round(hours_remaining, 2),
        "ttl_hours": INVENTORY_TTL_HOURS,
        "never_uploaded": False,
    }


async def require_fresh_inventory(current_user: dict = Depends(get_current_user)) -> dict:
    """Dependency that blocks mutating actions when inventory has expired."""
    status = await get_inventory_status()
    if not status["fresh"]:
        raise HTTPException(
            status_code=423,  # Locked
            detail={
                "code": "inventory_stale",
                "message": (
                    "Inventory has not been uploaded in the last "
                    f"{INVENTORY_TTL_HOURS} hours. Upload a fresh inventory file to continue."
                ),
                **status,
            },
        )
    return current_user


# ---------------------------------------------------------------------------
# Auth Endpoints
# ---------------------------------------------------------------------------
@api_router.post("/auth/login")
async def login(body: LoginRequest):
    user = await db.users.find_one({"username": body.username.lower()})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    token = create_access_token(user["id"], user["username"])
    role = user.get("role", "owner")
    systems = user.get("systems") or list(SYSTEMS)
    perms = default_permissions(all_true=(role == "owner"))
    perms.update(user.get("permissions") or {})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": user["id"],
            "username": user["username"],
            "role": role,
            "systems": systems,
            "permissions": perms,
        },
    }


@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return current_user


@api_router.put("/auth/change-credentials")
async def change_credentials(body: CredentialsUpdate, current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user["id"]})
    if not user or not verify_password(body.current_password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Current password is incorrect")

    updates: dict = {}
    new_username = (body.new_username or "").strip().lower()
    if new_username and new_username != user["username"]:
        if len(new_username) < 3:
            raise HTTPException(status_code=400, detail="Username must be at least 3 characters")
        existing = await db.users.find_one({"username": new_username, "id": {"$ne": user["id"]}})
        if existing:
            raise HTTPException(status_code=400, detail="Username is already taken")
        updates["username"] = new_username

    if body.new_password:
        if len(body.new_password) < 6:
            raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
        updates["password_hash"] = hash_password(body.new_password)

    if not updates:
        raise HTTPException(status_code=400, detail="Nothing to update")

    updates["updated_at"] = now_iso()
    await db.users.update_one({"id": user["id"]}, {"$set": updates})
    token = create_access_token(user["id"], updates.get("username", user["username"]))
    role = user.get("role", "owner")
    systems = user.get("systems") or list(SYSTEMS)
    perms = default_permissions(all_true=(role == "owner"))
    perms.update(user.get("permissions") or {})
    return {
        "success": True,
        "access_token": token,
        "user": {
            "id": user["id"],
            "username": updates.get("username", user["username"]),
            "role": role,
            "systems": systems,
            "permissions": perms,
        },
    }


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------
@api_router.get("/settings")
async def get_settings(current_user: dict = Depends(get_current_user)):
    settings = await db.settings.find_one({"key": "global"}, {"_id": 0})
    if not settings:
        settings = {"key": "global", "discount_percent": 0.0}
    return settings


@api_router.put("/settings/discount")
async def update_discount(body: DiscountUpdate, current_user: dict = Depends(get_current_user)):
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("change_discount"):
        raise HTTPException(status_code=403, detail="Missing permission: change_discount")
    if body.discount_percent < 0 or body.discount_percent > 100:
        raise HTTPException(status_code=400, detail="Discount must be between 0 and 100")
    await db.settings.update_one(
        {"key": "global"},
        {"$set": {"discount_percent": body.discount_percent, "updated_at": now_iso()}},
        upsert=True,
    )
    return {"discount_percent": body.discount_percent}


# ---------------------------------------------------------------------------
# Hero Search
# ---------------------------------------------------------------------------
@api_router.get("/hero/search")
async def hero_search(q: str, current_user: dict = Depends(require_fresh_inventory)):
    require_system_access(current_user, "hero")
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("search_ecatalogue"):
        raise HTTPException(status_code=403, detail="Missing permission: search_ecatalogue")
    q = (q or "").strip()
    if not q:
        raise HTTPException(status_code=400, detail="Search query required")
    try:
        data = hero_client.search_part(q)
    except requests.RequestException as e:
        logger.exception("hero search failed")
        raise HTTPException(status_code=502, detail=f"Hero eCatalogue unreachable: {e}")

    parts = []
    for p in data.get("searchParts") or []:
        raw_pn = p.get("partno") or p.get("sPartNo")
        parts.append({
            "part_no": format_part_no_display(raw_pn) if raw_pn else raw_pn,
            "part_no_original": raw_pn,
            "description": p.get("partdesc"),
            "type": p.get("type"),
            "moq": p.get("moq"),
            "mrp": p.get("mrp"),
            "image_url": None,
        })
    # attach image URL
    image_path = data.get("imagePath") or ""
    ecat_url = data.get("ecatUrl") or ""
    for p in parts:
        if p.get("part_no_original"):
            p["image_url"] = f"{ecat_url}{image_path}{p['part_no_original']}.jpg"
    return {"query": q, "parts": parts, "count": len(parts)}


@api_router.get("/tvs/search")
async def tvs_search(q: str, current_user: dict = Depends(require_fresh_inventory)):
    """Search the TVS advantagetvs.com eCatalogue by part number."""
    require_system_access(current_user, "tvs")
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("search_ecatalogue"):
        raise HTTPException(status_code=403, detail="Missing permission: search_ecatalogue")
    q = (q or "").strip()
    if not q:
        raise HTTPException(status_code=400, detail="Search query required")
    try:
        data = tvs_client.search_part(q)
    except requests.RequestException as e:
        logger.exception("tvs search failed")
        raise HTTPException(status_code=502, detail=f"TVS eCatalogue unreachable: {e}")

    inner = data.get("data") or {}
    rows = inner.get("ModelsAppies") or []

    # Dedupe per-partno keeping first (top) occurrence -- TVS returns one row per model variant.
    seen: set = set()
    parts: List[dict] = []
    for r in rows:
        pn = str(r.get("PARTNO") or "").strip()
        if not pn or pn in seen:
            continue
        seen.add(pn)
        try:
            mrp = float(r.get("MRP") or 0)
        except Exception:
            mrp = 0.0
        try:
            moq = float(r.get("MOQ") or 0)
        except Exception:
            moq = 0.0
        parts.append({
            "part_no": pn,
            "part_no_original": pn,
            "description": r.get("PartDescription") or "",
            "type": r.get("SEGMENT_NAME") or "",
            "series": r.get("SERIES_NAME") or "",
            "model": r.get("MODEL_NAME") or "",
            "variant": r.get("VARIENT") or "",
            "moq": moq,
            "mrp": mrp,
            "image_url": None,
        })
    return {"query": q, "parts": parts, "count": len(parts), "raw_rows": len(rows)}


# ---------------------------------------------------------------------------
# Orders
# ---------------------------------------------------------------------------
@api_router.get("/orders")
async def list_orders(
    system: str = "hero",
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    require_system_access(current_user, system)
    query: dict = {"system": system}
    if status in ("current", "sent"):
        query["status"] = status
    cursor = db.orders.find(query, {"_id": 0}).sort("created_at", -1)
    docs = await cursor.to_list(1000)
    return docs


@api_router.get("/orders/{order_id}")
async def get_order(order_id: str, current_user: dict = Depends(get_current_user)):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    require_system_access(current_user, order.get("system", "hero"))
    return order


async def _ensure_current_orders_limit(system: str) -> None:
    """Raise 409 if the concurrent-current-orders cap is already hit for this system."""
    current_count = await db.orders.count_documents({"status": "current", "system": system})
    if current_count >= MAX_CURRENT_ORDERS:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "current_orders_limit",
                "message": (
                    f"You already have {current_count} current {system.upper()} orders. "
                    f"Mark one as sent (or delete it) before starting a new one. "
                    f"Limit: {MAX_CURRENT_ORDERS}."
                ),
                "limit": MAX_CURRENT_ORDERS,
                "current_count": current_count,
                "system": system,
            },
        )


async def _resolve_new_order_items(
    incoming: List[OrderItem], global_discount: float, system: str = "hero"
) -> List[dict]:
    """Auto-inject mandatory parts (when toggle is on and incoming is empty),
    dedupe by normalized part number, and compute per-line totals."""
    mand_toggle = await db.settings.find_one({"key": f"mandatory_parts_toggle:{system}"}) or {}
    mand_enabled = bool(mand_toggle.get("enabled", False))

    items_in = list(incoming)
    if mand_enabled and not items_in:
        async for mp in db.mandatory_parts.find({"system": system}, {"_id": 0}):
            items_in.append(OrderItem(
                part_no=mp.get("part_no", ""),
                description=mp.get("description", ""),
                mrp=float(mp.get("mrp") or 0),
                qty=int(mp.get("qty") or 1),
            ))

    items: List[dict] = []
    seen: set = set()
    for it in items_in:
        d = it.model_dump()
        norm = normalize_part_no(d["part_no"])
        if not norm or norm in seen:
            continue  # dedupe silently on create
        seen.add(norm)
        compute_item_totals(d, global_discount)
        items.append(d)
    return items


@api_router.post("/orders")
async def create_order(
    body: OrderCreate,
    system: str = "hero",
    current_user: dict = Depends(require_fresh_inventory),
) -> dict:
    require_system_access(current_user, system)
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("orders_create_edit"):
        raise HTTPException(status_code=403, detail="Missing permission: orders_create_edit")
    await _ensure_current_orders_limit(system)

    settings = await db.settings.find_one({"key": "global"}) or {}
    global_discount = float(settings.get("discount_percent") or 0.0)
    items = await _resolve_new_order_items(body.items or [], global_discount, system)

    order = {
        "id": str(uuid.uuid4()),
        "order_no": await generate_order_no(system),
        "system": system,
        "status": "current",
        "items": items,
        "remarks": body.remarks,
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "sent_at": None,
        "created_by": current_user["username"],
        "global_discount_snapshot": global_discount,
    }
    await db.orders.insert_one(order)
    order.pop("_id", None)
    return order


@api_router.put("/orders/{order_id}")
async def update_order(order_id: str, body: OrderUpdate, current_user: dict = Depends(require_fresh_inventory)):
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    require_system_access(current_user, order.get("system", "hero"))
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("orders_create_edit"):
        raise HTTPException(status_code=403, detail="Missing permission: orders_create_edit")
    if order.get("status") == "sent":
        raise HTTPException(status_code=400, detail="Cannot edit a sent order")

    settings = await db.settings.find_one({"key": "global"}) or {}
    global_discount = float(settings.get("discount_percent") or 0.0)

    items = []
    seen = set()
    for it in body.items:
        d = it.model_dump()
        norm = normalize_part_no(d["part_no"])
        if not norm:
            continue
        if norm in seen:
            raise HTTPException(status_code=400, detail=f"Duplicate part number: {d['part_no']}")
        seen.add(norm)
        compute_item_totals(d, global_discount)
        items.append(d)

    if not items:
        raise HTTPException(status_code=400, detail="Cannot save an empty order. Add at least one part.")

    await db.orders.update_one(
        {"id": order_id},
        {"$set": {"items": items, "remarks": body.remarks, "updated_at": now_iso()}},
    )
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
    return updated


@api_router.post("/orders/{order_id}/mark-sent")
async def mark_sent(order_id: str, current_user: dict = Depends(require_fresh_inventory)):
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    require_system_access(current_user, order.get("system", "hero"))
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("orders_mark_sent"):
        raise HTTPException(status_code=403, detail="Missing permission: orders_mark_sent")
    if not order.get("items"):
        raise HTTPException(status_code=400, detail="Cannot send an empty order")
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {"status": "sent", "sent_at": now_iso(), "updated_at": now_iso()}},
    )
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
    return updated


@api_router.post("/orders/{order_id}/reopen")
async def reopen_order(order_id: str, current_user: dict = Depends(get_current_user)):
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    require_system_access(current_user, order.get("system", "hero"))
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("orders_mark_sent"):
        raise HTTPException(status_code=403, detail="Missing permission: orders_mark_sent")
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {"status": "current", "sent_at": None, "updated_at": now_iso()}},
    )
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
    return updated


@api_router.delete("/orders/{order_id}")
async def delete_order(order_id: str, confirm: str = "", current_user: dict = Depends(get_current_user)):
    if confirm.strip().lower() != "delete":
        raise HTTPException(
            status_code=400,
            detail="Delete not confirmed. Type 'delete' to confirm.",
        )
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    require_system_access(current_user, order.get("system", "hero"))
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("orders_delete"):
        raise HTTPException(status_code=403, detail="Missing permission: orders_delete")
    await db.orders.delete_one({"id": order_id})
    return {"success": True}


@api_router.get("/orders/check-part/{part_no}")
async def check_part_history(
    part_no: str,
    system: str = "hero",
    exclude_order_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """Warn only if this part was in the MOST RECENT previous order sheet (same system)."""
    require_system_access(current_user, system)
    norm = normalize_part_no(part_no)
    query: dict = {"system": system, "items.part_no": {"$exists": True}}
    if exclude_order_id:
        query["id"] = {"$ne": exclude_order_id}
    # Find only the most recent order (excluding the current one)
    latest = await db.orders.find_one(
        query,
        {"_id": 0, "id": 1, "order_no": 1, "status": 1, "created_at": 1, "items": 1},
        sort=[("created_at", -1)],
    )
    prev_orders = []
    if latest:
        for it in latest.get("items", []):
            if normalize_part_no(it.get("part_no", "")) == norm:
                prev_orders.append({
                    "order_id": latest["id"],
                    "order_no": latest["order_no"],
                    "status": latest["status"],
                    "created_at": latest["created_at"],
                    "qty": it.get("qty"),
                })
                break
    return {"part_no": part_no, "previously_ordered": len(prev_orders) > 0, "orders": prev_orders}


# ---------------------------------------------------------------------------
# Excel & PDF Export
# ---------------------------------------------------------------------------
_EXCEL_HEADER_FILL = PatternFill("solid", fgColor="E31837")
_EXCEL_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_EXCEL_THIN = Side(border_style="thin", color="333333")
_EXCEL_BORDER = Border(left=_EXCEL_THIN, right=_EXCEL_THIN, top=_EXCEL_THIN, bottom=_EXCEL_THIN)
_EXCEL_COL_WIDTHS = [8, 24, 60, 10]
_EXCEL_ITEM_HEADERS = ["S.No.", "Part No.", "Description", "Qty"]


def _excel_add_logo(ws) -> None:
    if not LOGO_PATH.exists():
        return
    try:
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(str(LOGO_PATH))
        img.width = 60
        img.height = 60
        ws.add_image(img, "A1")
    except Exception:
        pass


def _system_brand_meta(system: str) -> Dict[str, str]:
    if system == "tvs":
        return {
            "brand_line": "TVS Motor Genuine Parts",
            "brand_color": "#1E3A8A",  # TVS blue
        }
    return {
        "brand_line": "Hero MotoCorp Genuine Parts Dealer",
        "brand_color": "#B31229",  # Hero red
    }


def _excel_write_header(ws, order: dict) -> None:
    ws.row_dimensions[1].height = 45
    ws["B1"] = "KABIR AUTO PARTS"
    meta = _system_brand_meta(order.get("system", "hero"))
    ws["B1"].font = Font(bold=True, size=18, color=meta["brand_color"].lstrip("#"))
    ws["B2"] = meta["brand_line"]
    ws["B2"].font = Font(italic=True, color="666666", size=10)
    ws.merge_cells("B1:E1")
    ws.merge_cells("B2:E2")

    ws["F1"] = "ORDER SHEET"
    ws["F1"].font = Font(bold=True, size=12)
    ws.merge_cells("F1:I1")
    ws["F2"] = order["order_no"]
    ws["F2"].font = Font(bold=True, size=11, color=meta["brand_color"].lstrip("#"))
    ws.merge_cells("F2:I2")

    ws["A4"] = f"Status: {order['status'].upper()}"
    ws["A5"] = f"Created: {order['created_at'][:19].replace('T', ' ')}"
    if order.get("sent_at"):
        ws["E4"] = f"Sent: {order['sent_at'][:19].replace('T', ' ')}"
    if order.get("remarks"):
        ws["A7"] = f"Remarks: {order['remarks']}"


def _excel_write_items_table(ws, order: dict, header_row: int = 9) -> int:
    """Write the items table starting at `header_row`. Returns total qty."""
    for i, h in enumerate(_EXCEL_ITEM_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.fill = _EXCEL_HEADER_FILL
        cell.font = _EXCEL_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _EXCEL_BORDER

    total_qty = 0
    items = order.get("items", []) or []
    is_hero = (order.get("system", "hero") == "hero")
    for idx, item in enumerate(items, start=1):
        row = header_row + idx
        pn_raw = item.get("part_no", "")
        pn_display = format_part_no_display(pn_raw) if is_hero else pn_raw
        vals = [
            idx,
            pn_display,
            item.get("description", ""),
            item.get("qty", 0),
        ]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = _EXCEL_BORDER
            if i in (1, 4):
                cell.alignment = Alignment(horizontal="center")
        total_qty += int(item.get("qty") or 0)

    total_row = header_row + len(items) + 1
    total_label = ws.cell(row=total_row, column=3, value="TOTAL QTY")
    total_label.font = Font(bold=True)
    total_label.alignment = Alignment(horizontal="right")
    total_qty_cell = ws.cell(row=total_row, column=4, value=total_qty)
    total_qty_cell.font = Font(bold=True)
    total_qty_cell.alignment = Alignment(horizontal="center")
    return total_qty


def build_order_excel(order: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    _excel_add_logo(ws)
    _excel_write_header(ws, order)
    _excel_write_items_table(ws, order)

    for i, w in enumerate(_EXCEL_COL_WIDTHS, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


LOGO_PATH = ROOT_DIR / "assets" / "kabir-logo.jpg"


def _pdf_styles():
    styles = getSampleStyleSheet()
    return {
        "brand": ParagraphStyle("brand", parent=styles["Heading1"], textColor=colors.HexColor("#B31229"), fontSize=20, spaceAfter=0),
        "subtitle": ParagraphStyle("subtitle", parent=styles["Normal"], textColor=colors.HexColor("#555555"), fontSize=10, spaceAfter=8),
        "meta": ParagraphStyle("meta", parent=styles["Normal"], fontSize=9),
    }


def _pdf_header_row(order: dict, style_map: dict) -> Table:
    logo_cell = ""
    if LOGO_PATH.exists():
        try:
            logo_cell = RLImage(str(LOGO_PATH), width=22 * mm, height=22 * mm)
        except Exception:
            logo_cell = ""
    meta = _system_brand_meta(order.get("system", "hero"))
    brand_para = Paragraph(
        "<b>KABIR AUTO PARTS</b><br/>"
        f"<font size=8 color='#666666'>{meta['brand_line']}</font>",
        style_map["brand"],
    )
    order_meta = Paragraph(
        f"<b>ORDER SHEET</b><br/>"
        f"<font size=10><b>{order['order_no']}</b></font><br/>"
        f"<font size=8 color='#666666'>{order['status'].upper()} · "
        f"{order['created_at'][:10]}</font>",
        style_map["subtitle"],
    )
    line_color = colors.HexColor(meta["brand_color"] if order.get("system") == "hero" else "#1E3A8A")
    header_table = Table(
        [[logo_cell, brand_para, order_meta]],
        colWidths=[26 * mm, 160 * mm, 80 * mm],
    )
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, -1), 1.2, line_color),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    return header_table


def _pdf_items_table(order: dict) -> Table:
    data = [["S.No.", "Part No.", "Description", "Qty"]]
    total_qty = 0
    is_hero = (order.get("system", "hero") == "hero")
    for idx, item in enumerate(order.get("items", []) or [], start=1):
        pn_raw = item.get("part_no", "")
        pn_display = format_part_no_display(pn_raw) if is_hero else pn_raw
        data.append([
            str(idx),
            pn_display,
            str(item.get("description", ""))[:80],
            str(item.get("qty", 0)),
        ])
        total_qty += int(item.get("qty") or 0)
    data.append(["", "", "TOTAL QTY", str(total_qty)])

    header_bg = colors.HexColor("#E31837" if is_hero else "#1E3A8A")
    table = Table(data, colWidths=[20 * mm, 45 * mm, 165 * mm, 25 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (3, 0), (3, -1), "CENTER"),
        ("ALIGN", (2, -1), (2, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F3F4F6")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return table


def _pdf_meta_paragraphs(order: dict, meta_style) -> List:
    meta_text = f"<b>Created:</b> {order['created_at'][:19].replace('T', ' ')}"
    if order.get("sent_at"):
        meta_text += f"   |   <b>Sent:</b> {order['sent_at'][:19].replace('T', ' ')}"
    paras: list = [Paragraph(meta_text, meta_style)]
    if order.get("remarks"):
        paras.append(Paragraph(f"<b>Remarks:</b> {order['remarks']}", meta_style))
    return paras


def build_order_pdf(order: dict) -> bytes:
    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out, pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )
    style_map = _pdf_styles()

    story: list = []
    story.append(_pdf_header_row(order, style_map))
    story.append(Spacer(1, 4 * mm))
    story.extend(_pdf_meta_paragraphs(order, style_map["meta"]))
    story.append(Spacer(1, 5 * mm))
    story.append(_pdf_items_table(order))
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(
        f"<font size=7 color='#888888'>Generated by Kabir Auto Parts · "
        f"{'Hero MotoCorp' if order.get('system','hero') == 'hero' else 'TVS Motor'} "
        f"Parts Ordering System</font>",
        style_map["subtitle"],
    ))

    doc.build(story)
    out.seek(0)
    return out.read()


@api_router.get("/orders/{order_id}/export/excel")
async def export_excel(order_id: str, current_user: dict = Depends(get_current_user)):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    content = build_order_excel(order)
    filename = f"{order['order_no']}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@api_router.get("/orders/{order_id}/export/pdf")
async def export_pdf(order_id: str, current_user: dict = Depends(get_current_user)):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    content = build_order_pdf(order)
    filename = f"{order['order_no']}.pdf"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# Inventory
# ---------------------------------------------------------------------------
@api_router.get("/inventory/mapping")
async def get_mapping(current_user: dict = Depends(get_current_user)):
    m = await db.settings.find_one({"key": "inventory_mapping"}, {"_id": 0})
    if not m:
        return {"part_no": "", "description": "", "stock_qty": "", "location": "", "rate": ""}
    return {k: m.get(k, "") for k in ("part_no", "description", "stock_qty", "location", "rate")}


@api_router.put("/inventory/mapping")
async def set_mapping(body: InventoryMapping, current_user: dict = Depends(get_current_user)):
    doc = body.model_dump()
    doc["key"] = "inventory_mapping"
    doc["updated_at"] = now_iso()
    await db.settings.update_one({"key": "inventory_mapping"}, {"$set": doc}, upsert=True)
    return {"success": True, "mapping": body.model_dump()}


def read_upload_to_df(file: UploadFile, content: bytes) -> pd.DataFrame:
    """Parse an uploaded CSV/XLSX file into a DataFrame. Raises HTTPException on error."""
    try:
        if file.filename.lower().endswith(".csv"):
            return pd.read_csv(io.BytesIO(content))
        return pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {e}")


@api_router.post("/inventory/preview")
async def inventory_preview(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Return the columns and first 5 rows of the uploaded Excel/CSV."""
    content = await file.read()
    df = read_upload_to_df(file, content)
    columns = [str(c) for c in df.columns.tolist()]
    sample = df.head(5).fillna("").astype(str).to_dict(orient="records")
    return {"columns": columns, "sample": sample, "row_count": int(len(df))}


def _parse_inventory_rows(df, mapping: dict) -> List[dict]:
    """Turn a DataFrame + column mapping into inventory documents ready to insert."""
    part_no_col = mapping["part_no"]
    stock_qty_col = mapping["stock_qty"]
    desc_col = mapping.get("description") or ""
    loc_col = mapping.get("location") or ""
    rate_col = mapping.get("rate") or ""

    ts = now_iso()
    rows: List[dict] = []
    for _, r in df.iterrows():
        raw_pn = r.get(part_no_col, "")
        if pd.isna(raw_pn) or str(raw_pn).strip() == "":
            continue
        try:
            qty_val = r.get(stock_qty_col, 0)
            qty_num = float(qty_val) if not pd.isna(qty_val) else 0.0
        except Exception:
            qty_num = 0.0
        rows.append({
            "id": str(uuid.uuid4()),
            "part_no": str(raw_pn).strip(),
            "part_no_norm": normalize_part_no(str(raw_pn)),
            "description": str(r.get(desc_col, "")).strip() if desc_col else "",
            "stock_qty": qty_num,
            "location": str(r.get(loc_col, "")).strip() if loc_col else "",
            "rate": str(r.get(rate_col, "")).strip() if rate_col else "",
            "uploaded_at": ts,
        })
    return rows


@api_router.post("/inventory/upload")
async def inventory_upload(
    file: UploadFile = File(...),
    part_no: str = Form(...),
    description: str = Form(""),
    stock_qty: str = Form(...),
    location: str = Form(""),
    rate: str = Form(""),
    replace: bool = Form(True),
    current_user: dict = Depends(get_current_user),
) -> dict:
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("inventory_upload"):
        raise HTTPException(status_code=403, detail="Missing permission: inventory_upload")
    content = await file.read()
    df = read_upload_to_df(file, content)

    if part_no not in df.columns or stock_qty not in df.columns:
        raise HTTPException(status_code=400, detail="Mapped columns not found in file")

    mapping = {
        "part_no": part_no,
        "description": description,
        "stock_qty": stock_qty,
        "location": location,
        "rate": rate,
    }

    # save mapping
    await db.settings.update_one(
        {"key": "inventory_mapping"},
        {"$set": {**mapping, "key": "inventory_mapping", "updated_at": now_iso()}},
        upsert=True,
    )

    if replace:
        await db.inventory.delete_many({})

    rows = _parse_inventory_rows(df, mapping)
    if rows:
        await db.inventory.insert_many(rows)

    # record last upload timestamp for 24h freshness tracking
    await db.settings.update_one(
        {"key": "inventory_status"},
        {"$set": {"key": "inventory_status", "last_uploaded_at": now_iso()}},
        upsert=True,
    )

    return {"success": True, "imported": len(rows), "replaced": replace}


@api_router.get("/inventory/status")
async def inventory_status(current_user: dict = Depends(get_current_user)):
    return await get_inventory_status()


@api_router.get("/inventory")
async def list_inventory(q: str = "", current_user: dict = Depends(get_current_user)):
    query: dict = {}
    if q:
        query["part_no_norm"] = {"$regex": normalize_part_no(q)}
    cursor = db.inventory.find(query, {"_id": 0}).limit(500)
    return await cursor.to_list(500)


@api_router.get("/inventory/lookup/{part_no}")
async def lookup_inventory(part_no: str, current_user: dict = Depends(get_current_user)):
    norm = normalize_part_no(part_no)
    doc = await db.inventory.find_one({"part_no_norm": norm}, {"_id": 0})
    if not doc:
        return {"found": False, "stock_qty": 0, "part_no": part_no}
    return {"found": True, **doc}


# ---------------------------------------------------------------------------
# Important Parts (low-stock alerts on dashboard)
# ---------------------------------------------------------------------------
@api_router.get("/important-parts")
async def list_important_parts(
    system: str = "hero",
    current_user: dict = Depends(get_current_user),
):
    require_system_access(current_user, system)
    docs = await db.important_parts.find({"system": system}, {"_id": 0}).sort("created_at", -1).to_list(500)
    # enrich with current stock
    for d in docs:
        inv = await db.inventory.find_one(
            {"part_no_norm": d.get("part_no_norm")}, {"_id": 0, "stock_qty": 1}
        )
        d["current_stock"] = float(inv["stock_qty"]) if inv else 0.0
        d["is_low"] = d["current_stock"] < float(d.get("threshold_qty") or 0)
    return docs


@api_router.post("/important-parts")
async def add_important_part(
    body: ImportantPartBody,
    system: str = "hero",
    current_user: dict = Depends(get_current_user),
):
    require_system_access(current_user, system)
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("manage_important_parts"):
        raise HTTPException(status_code=403, detail="Missing permission: manage_important_parts")
    pn = (body.part_no or "").strip()
    if not pn:
        raise HTTPException(status_code=400, detail="Part number is required")
    norm = normalize_part_no(pn)
    existing = await db.important_parts.find_one({"system": system, "part_no_norm": norm})
    if existing:
        raise HTTPException(status_code=400, detail=f"{pn} is already in the important list")
    display_pn = format_part_no_display(pn) if system == "hero" else pn
    doc = {
        "id": str(uuid.uuid4()),
        "system": system,
        "part_no": display_pn,
        "part_no_norm": norm,
        "description": body.description or "",
        "threshold_qty": float(body.threshold_qty or 1),
        "created_at": now_iso(),
    }
    await db.important_parts.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/important-parts/{item_id}")
async def update_important_part(item_id: str, body: ImportantPartBody, current_user: dict = Depends(get_current_user)):
    existing = await db.important_parts.find_one({"id": item_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Not found")
    require_system_access(current_user, existing.get("system", "hero"))
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("manage_important_parts"):
        raise HTTPException(status_code=403, detail="Missing permission: manage_important_parts")
    updates = {
        "description": body.description or "",
        "threshold_qty": float(body.threshold_qty or 1),
        "updated_at": now_iso(),
    }
    await db.important_parts.update_one({"id": item_id}, {"$set": updates})
    return {"success": True}


@api_router.delete("/important-parts/{item_id}")
async def delete_important_part(item_id: str, current_user: dict = Depends(get_current_user)):
    existing = await db.important_parts.find_one({"id": item_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Not found")
    require_system_access(current_user, existing.get("system", "hero"))
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("manage_important_parts"):
        raise HTTPException(status_code=403, detail="Missing permission: manage_important_parts")
    await db.important_parts.delete_one({"id": item_id})
    return {"success": True}


# ---------------------------------------------------------------------------
# Mandatory Parts (auto-add to every new order sheet when toggle is on)
# ---------------------------------------------------------------------------
@api_router.get("/mandatory-parts")
async def list_mandatory_parts(
    system: str = "hero",
    current_user: dict = Depends(get_current_user),
):
    require_system_access(current_user, system)
    docs = await db.mandatory_parts.find({"system": system}, {"_id": 0}).sort("created_at", -1).to_list(500)
    toggle = await db.settings.find_one({"key": f"mandatory_parts_toggle:{system}"}) or {}
    return {"parts": docs, "enabled": bool(toggle.get("enabled", False))}


@api_router.post("/mandatory-parts")
async def add_mandatory_part(
    body: MandatoryPartBody,
    system: str = "hero",
    current_user: dict = Depends(get_current_user),
):
    require_system_access(current_user, system)
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("manage_mandatory_parts"):
        raise HTTPException(status_code=403, detail="Missing permission: manage_mandatory_parts")
    pn = (body.part_no or "").strip()
    if not pn:
        raise HTTPException(status_code=400, detail="Part number is required")
    norm = normalize_part_no(pn)
    existing = await db.mandatory_parts.find_one({"system": system, "part_no_norm": norm})
    if existing:
        raise HTTPException(status_code=400, detail=f"{pn} is already in the mandatory list")
    display_pn = format_part_no_display(pn) if system == "hero" else pn
    doc = {
        "id": str(uuid.uuid4()),
        "system": system,
        "part_no": display_pn,
        "part_no_norm": norm,
        "description": body.description or "",
        "mrp": float(body.mrp or 0),
        "qty": int(body.qty or 1),
        "created_at": now_iso(),
    }
    await db.mandatory_parts.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/mandatory-parts/{item_id}")
async def update_mandatory_part(item_id: str, body: MandatoryPartBody, current_user: dict = Depends(get_current_user)):
    existing = await db.mandatory_parts.find_one({"id": item_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Not found")
    require_system_access(current_user, existing.get("system", "hero"))
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("manage_mandatory_parts"):
        raise HTTPException(status_code=403, detail="Missing permission: manage_mandatory_parts")
    updates = {
        "description": body.description or "",
        "mrp": float(body.mrp or 0),
        "qty": int(body.qty or 1),
        "updated_at": now_iso(),
    }
    await db.mandatory_parts.update_one({"id": item_id}, {"$set": updates})
    return {"success": True}


@api_router.delete("/mandatory-parts/{item_id}")
async def delete_mandatory_part(item_id: str, current_user: dict = Depends(get_current_user)):
    existing = await db.mandatory_parts.find_one({"id": item_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Not found")
    require_system_access(current_user, existing.get("system", "hero"))
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("manage_mandatory_parts"):
        raise HTTPException(status_code=403, detail="Missing permission: manage_mandatory_parts")
    await db.mandatory_parts.delete_one({"id": item_id})
    return {"success": True}


@api_router.put("/mandatory-toggle")
async def toggle_mandatory_parts(
    body: MandatoryToggleBody,
    system: str = "hero",
    current_user: dict = Depends(get_current_user),
):
    require_system_access(current_user, system)
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("manage_mandatory_parts"):
        raise HTTPException(status_code=403, detail="Missing permission: manage_mandatory_parts")
    key = f"mandatory_parts_toggle:{system}"
    await db.settings.update_one(
        {"key": key},
        {"$set": {"key": key, "enabled": bool(body.enabled), "updated_at": now_iso()}},
        upsert=True,
    )
    return {"enabled": bool(body.enabled)}


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
@api_router.get("/dashboard/stats")
async def dashboard_stats(system: str = "hero", current_user: dict = Depends(get_current_user)):
    require_system_access(current_user, system)
    current = await db.orders.count_documents({"status": "current", "system": system})
    sent = await db.orders.count_documents({"status": "sent", "system": system})
    inventory_count = await db.inventory.count_documents({})
    total_sent_value = 0.0
    async for o in db.orders.find({"status": "sent", "system": system}, {"items": 1, "_id": 0}):
        for it in o.get("items", []):
            total_sent_value += float(it.get("line_total") or 0)

    # low-stock alerts on important parts (this system only)
    low_stock_alerts = []
    async for ip in db.important_parts.find({"system": system}, {"_id": 0}):
        inv = await db.inventory.find_one(
            {"part_no_norm": ip.get("part_no_norm")}, {"_id": 0, "stock_qty": 1}
        )
        current_stock = float(inv["stock_qty"]) if inv else 0.0
        threshold = float(ip.get("threshold_qty") or 0)
        if current_stock < threshold:
            low_stock_alerts.append({
                "id": ip["id"],
                "part_no": ip["part_no"],
                "description": ip.get("description", ""),
                "threshold_qty": threshold,
                "current_stock": current_stock,
            })

    inv_status = await get_inventory_status()

    return {
        "system": system,
        "current_orders": current,
        "sent_orders": sent,
        "inventory_items": inventory_count,
        "total_sent_value": round(total_sent_value, 2),
        "low_stock_alerts": low_stock_alerts,
        "inventory_status": inv_status,
        "current_orders_limit": MAX_CURRENT_ORDERS,
        "current_orders_full": current >= MAX_CURRENT_ORDERS,
    }


# ---------------------------------------------------------------------------
# Employees (owner-only user management)
# ---------------------------------------------------------------------------
def _sanitize_permissions(perms: Optional[Dict[str, bool]]) -> Dict[str, bool]:
    result = default_permissions(all_true=False)
    if perms:
        for k, v in perms.items():
            if k in result:
                result[k] = bool(v)
    return result


def _sanitize_systems(systems: Optional[List[str]]) -> List[str]:
    out: List[str] = []
    for s in (systems or []):
        if s in SYSTEMS and s not in out:
            out.append(s)
    if not out:
        out = ["hero"]
    return out


def _public_user(user: dict) -> dict:
    return {
        "id": user.get("id"),
        "username": user.get("username"),
        "role": user.get("role", "employee"),
        "systems": user.get("systems") or [],
        "permissions": user.get("permissions") or default_permissions(all_true=(user.get("role") == "owner")),
        "created_at": user.get("created_at"),
        "updated_at": user.get("updated_at"),
    }


@api_router.get("/permissions/keys")
async def list_permission_keys(current_user: dict = Depends(get_current_user)):
    """Returns the list of permission keys with human-readable labels."""
    return {
        "keys": [
            {"key": "orders_create_edit", "label": "Create / edit orders"},
            {"key": "orders_delete", "label": "Delete orders"},
            {"key": "orders_mark_sent", "label": "Mark orders sent / reopen"},
            {"key": "search_ecatalogue", "label": "Search eCatalogue (Hero / TVS)"},
            {"key": "inventory_view", "label": "View inventory"},
            {"key": "inventory_upload", "label": "Upload / replace inventory"},
            {"key": "manage_important_parts", "label": "Manage important parts"},
            {"key": "manage_mandatory_parts", "label": "Manage mandatory parts"},
            {"key": "change_discount", "label": "Change discount setting"},
            {"key": "backup_restore", "label": "Export / import database backup"},
        ],
        "systems": list(SYSTEMS),
    }


@api_router.get("/employees")
async def list_employees(current_user: dict = Depends(require_owner)):
    docs = await db.users.find(
        {"role": {"$ne": "owner"}},
        {"_id": 0, "password_hash": 0},
    ).sort("created_at", -1).to_list(500)
    return [_public_user(d) for d in docs]


@api_router.post("/employees")
async def create_employee(body: EmployeeCreate, current_user: dict = Depends(require_owner)):
    uname = (body.username or "").strip().lower()
    if len(uname) < 3:
        raise HTTPException(status_code=400, detail="Username must be at least 3 characters")
    if len(body.password or "") < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    if await db.users.find_one({"username": uname}):
        raise HTTPException(status_code=400, detail="Username is already taken")

    doc = {
        "id": str(uuid.uuid4()),
        "username": uname,
        "password_hash": hash_password(body.password),
        "role": "employee",
        "systems": _sanitize_systems(body.systems),
        "permissions": _sanitize_permissions(body.permissions),
        "created_at": now_iso(),
    }
    await db.users.insert_one(doc)
    return _public_user(doc)


@api_router.put("/employees/{employee_id}")
async def update_employee(employee_id: str, body: EmployeeUpdate, current_user: dict = Depends(require_owner)):
    user = await db.users.find_one({"id": employee_id})
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    if user.get("role") == "owner":
        raise HTTPException(status_code=400, detail="Cannot modify owner via employees endpoint")

    updates: dict = {"updated_at": now_iso()}
    if body.password:
        if len(body.password) < 6:
            raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
        updates["password_hash"] = hash_password(body.password)
    if body.systems is not None:
        updates["systems"] = _sanitize_systems(body.systems)
    if body.permissions is not None:
        updates["permissions"] = _sanitize_permissions(body.permissions)

    await db.users.update_one({"id": employee_id}, {"$set": updates})
    fresh = await db.users.find_one({"id": employee_id}, {"_id": 0, "password_hash": 0})
    return _public_user(fresh or {})


@api_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str, current_user: dict = Depends(require_owner)):
    user = await db.users.find_one({"id": employee_id})
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    if user.get("role") == "owner":
        raise HTTPException(status_code=400, detail="Cannot delete owner account")
    await db.users.delete_one({"id": employee_id})
    return {"success": True}


# ---------------------------------------------------------------------------
# Database Export / Import (full backup)
# ---------------------------------------------------------------------------
# Collections we back up. `users` is included so credentials survive a restore.
BACKUP_COLLECTIONS = [
    "users",
    "settings",
    "counters",
    "orders",
    "inventory",
    "important_parts",
    "mandatory_parts",
]


@api_router.get("/db/export")
async def export_database(current_user: dict = Depends(get_current_user)):
    """Stream a JSON backup of the entire database."""
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("backup_restore"):
        raise HTTPException(status_code=403, detail="Missing permission: backup_restore")
    import json as _json

    payload: Dict[str, Any] = {
        "app": "hero-parts-ordering",
        "version": 1,
        "exported_at": now_iso(),
        "exported_by": current_user.get("username"),
        "collections": {},
    }
    for name in BACKUP_COLLECTIONS:
        docs = await db[name].find({}, {"_id": 0}).to_list(None)
        payload["collections"][name] = docs

    body = _json.dumps(payload, indent=2, default=str).encode("utf-8")
    filename = f"hmcl-backup-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.json"
    return StreamingResponse(
        io.BytesIO(body),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _parse_backup_payload(raw: bytes) -> dict:
    import json as _json
    try:
        data = _json.loads(raw.decode("utf-8"))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Invalid JSON file: {exc}")

    if not isinstance(data, dict) or data.get("app") != "hero-parts-ordering":
        raise HTTPException(
            status_code=400,
            detail="Not a Hero Parts Ordering backup file.",
        )
    if not isinstance(data.get("collections"), dict):
        raise HTTPException(status_code=400, detail="Missing 'collections' object in backup.")
    return data


async def _replace_collection_from_backup(name: str, docs: list) -> None:
    if not isinstance(docs, list):
        raise HTTPException(status_code=400, detail=f"Collection '{name}' must be an array.")
    await db[name].drop()
    if docs:
        # Strip any Mongo _id from source to avoid conflicts
        for d in docs:
            if isinstance(d, dict):
                d.pop("_id", None)
        await db[name].insert_many(docs)


async def _recreate_core_indexes() -> None:
    """Mirror index creation done in on_startup (called after import overwrites)."""
    await db.users.create_index("username", unique=True)
    await db.orders.create_index("order_no", unique=True)
    await db.orders.create_index("status")
    await db.orders.create_index("system")
    await db.inventory.create_index("part_no_norm")
    # important_parts and mandatory_parts are unique per (system, part_no_norm)
    try:
        await db.important_parts.drop_index("part_no_norm_1")
    except Exception:
        pass
    try:
        await db.mandatory_parts.drop_index("part_no_norm_1")
    except Exception:
        pass
    await db.important_parts.create_index(
        [("system", 1), ("part_no_norm", 1)], unique=True, name="system_part_no_norm_uniq"
    )
    await db.mandatory_parts.create_index(
        [("system", 1), ("part_no_norm", 1)], unique=True, name="system_part_no_norm_uniq"
    )


@api_router.post("/db/import")
async def import_database(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
) -> dict:
    """Restore the database from a previously-exported JSON file. Wipes existing
    data in the affected collections and re-inserts the file's contents."""
    if not is_owner(current_user) and not current_user.get("permissions", {}).get("backup_restore"):
        raise HTTPException(status_code=403, detail="Missing permission: backup_restore")
    raw = await file.read()
    data = _parse_backup_payload(raw)
    collections = data["collections"]

    imported: dict = {}
    for name in BACKUP_COLLECTIONS:
        if name not in collections:
            continue
        docs = collections.get(name) or []
        await _replace_collection_from_backup(name, docs)
        imported[name] = len(docs)

    await _recreate_core_indexes()

    return {
        "success": True,
        "imported": imported,
        "exported_at": data.get("exported_at"),
        "note": "Sign out and sign in again if your credentials came from the imported file.",
    }


# ---------------------------------------------------------------------------
# Root & Startup
# ---------------------------------------------------------------------------
@api_router.get("/")
async def root():
    return {"service": "Hero Parts Ordering", "status": "ok"}


app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def on_startup():
    # Backfill 'system' field on legacy documents (before creating new indexes).
    await db.orders.update_many(
        {"system": {"$exists": False}}, {"$set": {"system": "hero"}}
    )
    await db.important_parts.update_many(
        {"system": {"$exists": False}}, {"$set": {"system": "hero"}}
    )
    await db.mandatory_parts.update_many(
        {"system": {"$exists": False}}, {"$set": {"system": "hero"}}
    )
    # Migrate old mandatory_parts_toggle setting key -> mandatory_parts_toggle:hero
    old_toggle = await db.settings.find_one({"key": "mandatory_parts_toggle"})
    if old_toggle:
        await db.settings.update_one(
            {"key": "mandatory_parts_toggle:hero"},
            {"$set": {
                "key": "mandatory_parts_toggle:hero",
                "enabled": bool(old_toggle.get("enabled", False)),
                "updated_at": now_iso(),
            }},
            upsert=True,
        )
        await db.settings.delete_one({"key": "mandatory_parts_toggle"})

    # indexes (may need to drop old ones)
    await db.users.create_index("username", unique=True)
    await db.orders.create_index("order_no", unique=True)
    await db.orders.create_index("status")
    await db.orders.create_index("system")
    await db.inventory.create_index("part_no_norm")
    # Drop legacy single-field unique indexes if present -- replaced by composite
    for coll in (db.important_parts, db.mandatory_parts):
        try:
            await coll.drop_index("part_no_norm_1")
        except Exception:
            pass
    await db.important_parts.create_index(
        [("system", 1), ("part_no_norm", 1)], unique=True, name="system_part_no_norm_uniq"
    )
    await db.mandatory_parts.create_index(
        [("system", 1), ("part_no_norm", 1)], unique=True, name="system_part_no_norm_uniq"
    )

    # seed admin only if no owner user exists
    admin_username = os.environ.get("ADMIN_USERNAME", "admin").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    any_owner = await db.users.find_one({"role": "owner"})
    if any_owner is None:
        # Upgrade any pre-existing legacy admin user, else create fresh owner
        legacy = await db.users.find_one({"username": admin_username})
        if legacy is not None:
            await db.users.update_one(
                {"id": legacy["id"]},
                {"$set": {
                    "role": "owner",
                    "systems": list(SYSTEMS),
                    "permissions": default_permissions(all_true=True),
                    "updated_at": now_iso(),
                }},
            )
            logger.info(f"Upgraded existing user '{admin_username}' to owner")
        else:
            await db.users.insert_one({
                "id": str(uuid.uuid4()),
                "username": admin_username,
                "password_hash": hash_password(admin_password),
                "role": "owner",
                "systems": list(SYSTEMS),
                "permissions": default_permissions(all_true=True),
                "created_at": now_iso(),
            })
            logger.info(f"Seeded owner user: {admin_username}")

    # seed default settings
    if not await db.settings.find_one({"key": "global"}):
        await db.settings.insert_one({"key": "global", "discount_percent": 25.0, "updated_at": now_iso()})


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
