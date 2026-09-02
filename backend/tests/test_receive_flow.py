"""Tests for the 'Mark as Received' feature (receive-check, mark-received, pending exports)."""
import io
import os
import re
from pathlib import Path

import pytest
import requests
from dotenv import dotenv_values

frontend_env = dotenv_values("/app/frontend/.env")
base_url = os.environ.get("REACT_APP_BACKEND_URL") or frontend_env.get("REACT_APP_BACKEND_URL")
if not base_url:
    raise RuntimeError("REACT_APP_BACKEND_URL missing")
BASE_URL = base_url.rstrip("/")

SENT_ORDER_ID = "996cf88d-1c2d-40f3-b1d9-0feadef9af6b"  # legacy sent hero order (no stock_at_sent)


@pytest.fixture(scope="module")
def creds():
    p = Path("/app/memory/test_credentials.md")
    if not p.exists():
        pytest.skip("missing test credentials file")
    c = p.read_text()
    u = re.search(r"(?im)^\s*[-*]?\s*(?:\*\*)?Username(?:\*\*)?\s*:\s*`?([^`\s]+)", c)
    pw = re.search(r"(?im)^\s*[-*]?\s*(?:\*\*)?Password(?:\*\*)?\s*:\s*`?([^`\s]+)", c)
    if not u or not pw:
        pytest.skip("no creds parsed")
    return {"username": u.group(1), "password": pw.group(1)}


@pytest.fixture(scope="module")
def client(creds):
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{BASE_URL}/api/auth/login", json=creds, timeout=30)
    if r.status_code != 200:
        pytest.fail(f"login failed {r.status_code}: {r.text[:300]}")
    tok = r.json().get("access_token")
    if not tok:
        pytest.fail(f"no access_token in login response: {r.text[:300]}")
    s.headers.update({"Authorization": f"Bearer {tok}"})
    return s


# ---------------- receive-check (legacy order, no snapshot) ----------------
class TestReceiveCheck:
    def test_receive_check_on_sent_order(self, client):
        r = client.get(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/receive-check")
        assert r.status_code == 200, r.text[:300]
        d = r.json()
        assert d["order_id"] == SENT_ORDER_ID
        assert isinstance(d["order_no"], str) and d["order_no"]
        assert "has_snapshot" in d and "existing_receipt" in d
        assert len(d["items"]) > 0
        for it in d["items"]:
            for k in ("part_no", "qty", "stock_at_sent", "current_stock",
                      "received_qty", "pending_qty", "status", "method"):
                assert k in it, f"missing {k}"
            assert it["status"] in ("received", "partial", "not_received")
            assert it["method"] in ("snapshot", "stock_level")
            assert it["pending_qty"] == max(0, it["qty"] - it["received_qty"])
        # legacy order -> stock_level method
        assert d["has_snapshot"] is False
        assert all(i["method"] == "stock_level" for i in d["items"])

    def test_receive_check_on_current_order_returns_400(self, client):
        r = client.get(f"{BASE_URL}/api/orders?system=hero&status=current")
        assert r.status_code == 200, r.text[:200]
        body = r.json()
        orders = body if isinstance(body, list) else body.get("orders", [])
        if not orders:
            pytest.skip("no current hero order available")
        oid = orders[0]["id"]
        rc = client.get(f"{BASE_URL}/api/orders/{oid}/receive-check")
        assert rc.status_code == 400, f"expected 400, got {rc.status_code}: {rc.text[:200]}"

    def test_receive_check_unknown_order_404(self, client):
        r = client.get(f"{BASE_URL}/api/orders/does-not-exist-xyz/receive-check")
        assert r.status_code == 404


# ---------------- mark-received overrides + receipt + exports ----------------
class TestMarkReceived:
    @pytest.fixture(scope="class", autouse=True)
    def restore_receipt(self, client):
        pre = client.get(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/receive-check").json()
        yield pre
        client.post(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/clear-receipt", json={})

    def test_mark_received_with_overrides(self, client, restore_receipt):
        items = restore_receipt["items"]
        assert len(items) >= 3
        p_not, p_partial, p_full = items[0], items[1], items[2]
        payload = {"items": [
            {"part_no": p_not["part_no"], "received": False},
            {"part_no": p_partial["part_no"], "received": True, "received_qty": 1},
            {"part_no": p_full["part_no"], "received": True},
        ]}
        r = client.post(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/mark-received", json=payload)
        assert r.status_code == 200, r.text[:400]
        order = r.json()
        assert "_id" not in order
        assert order["status"] == "sent", "order status must remain 'sent'"
        rec = order["receipt"]
        for k in ("received_at", "received_by", "items", "received_count", "partial_count", "pending_count"):
            assert k in rec
        by_pn = {i["part_no"]: i for i in rec["items"]}
        assert by_pn[p_not["part_no"]]["status"] == "not_received"
        assert by_pn[p_not["part_no"]]["pending_qty"] == p_not["qty"]
        pp = by_pn[p_partial["part_no"]]
        if p_partial["qty"] > 1:
            assert pp["status"] == "partial", pp
            assert pp["received_qty"] == 1
            assert pp["pending_qty"] == p_partial["qty"] - 1
        pf = by_pn[p_full["part_no"]]
        assert pf["status"] == "received" and pf["pending_qty"] == 0
        # counts consistent
        assert rec["received_count"] == sum(1 for i in rec["items"] if i["status"] == "received")
        assert rec["pending_count"] == sum(1 for i in rec["items"] if i["status"] != "received")
        # unlisted parts keep auto result
        for it in items[3:]:
            got = by_pn[it["part_no"]]
            assert got["received_qty"] == it["received_qty"], got
            assert got["method"] == it["method"]

    def test_receipt_persisted_and_visible_in_receive_check(self, client):
        g = client.get(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}")
        assert g.status_code == 200
        o = g.json()
        o = o.get("order", o)
        assert o.get("receipt") is not None
        assert o.get("received_at")
        rc = client.get(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/receive-check").json()
        assert rc["existing_receipt"] is not None

    def test_export_pending_excel(self, client):
        r = client.get(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/export-pending/excel")
        assert r.status_code == 200, r.text[:300]
        assert r.content[:2] == b"PK"
        assert "-PENDING.xlsx" in r.headers.get("content-disposition", "")
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl unavailable")
        wb = load_workbook(io.BytesIO(r.content))
        text = "\n".join(
            str(c.value) for row in wb.active.iter_rows() for c in row if c.value is not None
        )
        assert "-PENDING" in text
        # pending parts present, fully-received parts absent
        rec = client.get(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/receive-check").json()["existing_receipt"]
        pending = [i for i in rec["items"] if i["pending_qty"] > 0]
        received = [i for i in rec["items"] if i["pending_qty"] == 0]
        for i in pending:
            assert i["part_no"] in text, f"pending part {i['part_no']} missing from excel"
        for i in received:
            assert i["part_no"] not in text, f"received part {i['part_no']} should not be in excel"

    def test_export_pending_pdf(self, client):
        r = client.get(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/export-pending/pdf")
        assert r.status_code == 200, r.text[:300]
        assert r.content[:4] == b"%PDF"
        assert "-PENDING.pdf" in r.headers.get("content-disposition", "")

    def test_export_pending_unknown_kind_404(self, client):
        r = client.get(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/export-pending/csv")
        assert r.status_code == 404, r.text[:200]

    def test_mark_received_again_overwrites_receipt(self, client, restore_receipt):
        first = client.get(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/receive-check").json()["existing_receipt"]
        payload = {"items": [{"part_no": i["part_no"], "received": True} for i in restore_receipt["items"]]}
        r = client.post(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/mark-received", json=payload)
        assert r.status_code == 200, r.text[:300]
        rec = r.json()["receipt"]
        assert rec["pending_count"] == 0
        assert rec["received_count"] == len(restore_receipt["items"])
        assert rec["received_at"] != first["received_at"] or rec["pending_count"] != first["pending_count"]

    def test_export_pending_400_when_nothing_pending(self, client):
        # previous test marked everything received
        for kind in ("excel", "pdf"):
            r = client.get(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/export-pending/{kind}")
            assert r.status_code == 400, f"{kind}: {r.status_code} {r.text[:200]}"

    def test_clear_receipt_then_export_pending_400(self, client):
        r = client.post(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/clear-receipt", json={})
        assert r.status_code == 200, r.text[:300]
        o = r.json()
        assert o.get("receipt") is None
        assert o.get("received_at") is None
        assert o["status"] == "sent"
        e = client.get(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/export-pending/excel")
        assert e.status_code == 400
        assert "not been marked" in e.json().get("detail", "").lower()

    def test_full_order_export_regression(self, client):
        x = client.get(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/export/excel")
        assert x.status_code == 200 and x.content[:2] == b"PK"
        p = client.get(f"{BASE_URL}/api/orders/{SENT_ORDER_ID}/export/pdf")
        assert p.status_code == 200 and p.content[:4] == b"%PDF"


# ---------------- snapshot flow on a fresh order (tvs to avoid hero conflicts) ----------------
class TestStockSnapshot:
    order_id = None

    @pytest.fixture(scope="class", autouse=True)
    def cleanup(self, client):
        yield
        if TestStockSnapshot.order_id:
            client.delete(f"{BASE_URL}/api/orders/{TestStockSnapshot.order_id}?confirm=delete")

    def test_mark_sent_stores_stock_at_sent_and_snapshot_check(self, client):
        part = "91255GAA003RS"
        inv = client.get(f"{BASE_URL}/api/inventory/search?q={part}")
        assert inv.status_code == 200, inv.text[:200]
        body = inv.json()
        results = body if isinstance(body, list) else body.get("results", body.get("items", []))
        if not results:
            pytest.skip(f"part {part} not in inventory")
        stock = float(results[0].get("stock_qty") or 0)

        create = client.post(
            f"{BASE_URL}/api/orders?system=tvs",
            json={"items": [{"part_no": part, "qty": 2}], "remarks": "TEST_receive_snapshot"},
        )
        assert create.status_code in (200, 201), create.text[:400]
        order = create.json()
        order = order.get("order", order)
        TestStockSnapshot.order_id = order["id"]

        sent = client.post(f"{BASE_URL}/api/orders/{order['id']}/mark-sent", json={})
        assert sent.status_code == 200, sent.text[:400]
        so = sent.json()
        so = so.get("order", so)
        assert so["status"] == "sent"
        assert so["items"][0]["stock_at_sent"] == stock, so["items"][0]

        rc = client.get(f"{BASE_URL}/api/orders/{order['id']}/receive-check")
        assert rc.status_code == 200, rc.text[:300]
        d = rc.json()
        assert d["has_snapshot"] is True
        row = d["items"][0]
        assert row["method"] == "snapshot"
        assert row["received_qty"] == 0
        assert row["status"] == "not_received"
        assert row["pending_qty"] == row["qty"]

    def test_reopen_clears_receipt(self, client):
        oid = TestStockSnapshot.order_id
        if not oid:
            pytest.skip("no test order")
        mr = client.post(f"{BASE_URL}/api/orders/{oid}/mark-received",
                         json={"items": [{"part_no": "91255GAA003RS", "received": True, "received_qty": 1}]})
        assert mr.status_code == 200, mr.text[:300]
        assert mr.json()["receipt"]["partial_count"] == 1

        ro = client.post(f"{BASE_URL}/api/orders/{oid}/reopen", json={})
        assert ro.status_code == 200, ro.text[:300]
        o = ro.json()
        o = o.get("order", o)
        assert o["status"] == "current"
        assert o.get("receipt") is None
        assert o.get("received_at") is None
